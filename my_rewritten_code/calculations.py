"""
Name: .py

Purpose:

Functions:

Side effects:

"""

# Standard library imports
import datetime as datetime
import os
import sys
from typing import Any

# Third-party library imports
import numpy as np
import openpyxl
import pandas as pd
# Project-specific imports
from equitable.db.psyw import SmartDB
from equitable.infrastructure import sysenv
from scipy import interpolate

# Required custom modules
import file_utils

# Pandas configuration
pd.set_option('display.width', 150)

# Add system paths
sys.path.append(sysenv.get("ALM_DIR"))  # Add ALM_DIR to system path for additional modules

# Database connections
BM_conn = SmartDB('Benchmark')
BM_cur = BM_conn.con.cursor()

Bond_conn = SmartDB('Bond')
Bond_cur = Bond_conn.con.cursor()

General_conn = SmartDB('General')
General_cur = General_conn.con.cursor()

### Reads in input ###
"""Reads in assets"""
def get_expected_returns() -> pd.DataFrame:
    """
    Reads and interpolates expected bond returns from the "Parallel_tilt_curve_history.xlsx" file for various bond
    ratings and maturity terms.

    Returns:
    pd.DataFrame: A DataFrame containing the interpolated expected returns for different bond ratings and term assumptions.
    """
    file_name = "Parallel_tilt_curve_history.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    expected_returns = pd.DataFrame()
    ratings = ['Federal', 'Provincial', 'corporateAAA_AA', 'corporateA', 'corporateBBB']

    # Read the expected return data for each bond rating
    for sheet in ['analysis_quarterly_RF', 'analysis_quarterly_prov', 'analysis_quarterly_AA', 'analysis_quarterly_A', 'analysis_quarterly_BBB']:
        rownum = 27 if sheet == 'analysis_quarterly_RF' else 22
        ws = workbook[sheet]
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        returns = df.loc[rownum:rownum, 'term1':'term30']
        expected_returns = pd.concat([expected_returns, returns], ignore_index=True)

    # Assign bond ratings to the expected returns DataFrame
    expected_returns['ratings'] = ratings
    expected_returns.set_index('ratings', inplace=True)

    # Term assumptions for interpolation
    term_assumptions = [2, 7, 12, 17, 23, 29]
    return_assumptions = pd.DataFrame(columns=[0, 1, 2, 3, 4, 5])

    x = [1, 2, 3, 4, 5, 7, 10, 20, 30]
    # Interpolate expected returns for each rating across terms
    for rating in ratings:
        y = expected_returns.loc[rating].to_numpy()
        temp = interpolate.splrep(x, y, s=0)
        xnew = np.arange(1, 31)
        ynew = interpolate.splev(xnew, temp, der=0)
        return_assumptions.loc[rating] = ynew[term_assumptions]

    return return_assumptions / 100  # Convert to percentage returns




def BSTotals(given_date: datetime, sheet_version: int) -> dict:
    """
    Retrieves the balance sheet totals from the "SBS Totals.xlsx" file based on the provided date.

    Parameters:
    given_date (datetime): The date for which the balance sheet totals are requested.
    sheet_version (int): Determines if totals or segments are returned (1 for segments, 0 for totals).

    Returns:
    dict: A dictionary containing balance sheet totals for different categories.
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1

    year_quarter = f"{year}Q{quarter}"

    quarter = f"Q{quarter}"

    # file_name = "SBS Totals Modified.xlsx"
    # file_name = "SBS Totals - Brenda.xlsx"
    file_name = "SBS Totals.xlsx"
    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True)

    # Retrieve the appropriate worksheet based on sheet version
    # ws = workbook[year_quarter] if sheet_version == 1 else workbook[year_quarter + ' (Total)']
    ws = workbook['Segments'] if sheet_version == 1 else workbook['Total']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # Extract and return totals for relevant categories
    totals = {
        'ACCUM': df.loc[2, 'ACCUM'],
        'PAYOUT': df.loc[2, 'PAYOUT'],
        'UNIVERSAL': df.loc[2, 'UNIVERSAL'],
        'NONPAR': df.loc[2, 'NONPAR'],
        'GROUP': df.loc[2, 'GROUP'],
        'PARCSM': df.loc[2, 'PARCSM'],
        'SEGFUNDS': df.loc[2, 'SEGFUNDS'],
        'Surplus': df.loc[2, 'Surplus'],
        'Total': df.loc[2, 'Total']
    } # rewrote BSTotals to reduce unecessary code / overwritten initializaitons

    return totals

# Asset mix.xlsx reading in function
def percents(given_date: datetime, curMonthBS: bool = False) -> pd.DataFrame:
    """
    Retrieves asset mix percentages from the "Asset Mix.xlsx" file for the given date.

    Parameters:
    given_date (datetime): The date for which the asset mix percentages are requested.
    curMonthBS (bool): If True, adjusts the quarter to the next one if applicable. Default is False.

    Returns:
    pd.DataFrame: A DataFrame containing asset mix percentages for various bond ratings.
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    if curMonthBS and quarter < 4:
        quarter += 1
    year_quarter = year + "Q" + str(quarter)

    quarter = f"Q{quarter}"

    # file_name = "Asset Mix Modified.xlsx"
    file_name = "Asset Mix.xlsx"
    # file_name = "Asset Mix Prov -1%.xlsx" # TODO! for run2

    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    workbook = openpyxl.load_workbook(path_input, data_only = True) # TODO! new change - linked data vals work
    # ws = workbook[year_quarter]  # i.e., ws = workbook['2024Q1']
    ws = workbook['Sheet1']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns).set_index('rating')

    # Initialize surplus and SEGFUNDS columns
    df['Surplus'] = df['SEGFUNDS'] = 0

    # Filter rows to include only relevant bond categories
    df = df.loc[['Federal',
                 'Provincial',
                 'CorpAAA_AA',
                 'CorpA',
                 'CorpBBB',
                 'MortgagesInsured',
                 'MortgagesConv',
                 'PrivateAA',
                 'PrivateA',
                 'PrivateBBB',
                 'PrivateBB_B']]

    return df
# end of Asset Mix.xlsx reading in funciton
from typing import Dict

def solution_dollar_amounts(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculates the dollar allocation for each bond rating across different portfolios based on the asset mix and solution data.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations by rating.

    Returns:
    pd.DataFrame: A DataFrame with dollar allocations for each portfolio and rating.
    """
    weights = Asset_mix[['Accum', 'group', 'ul', 'Payout', 'np']].stack().sort_index()  # (*Brenda)
    weights2 = weights.reset_index(drop=True)

    # Filter the solution DataFrame to exclude 'Liability' and 'Total' portfolios
    sols = solution_df[(solution_df['portfolio'] != 'Liability') & (solution_df['portfolio'] != 'Total')].set_index \
        (['rating', 'portfolio']).sort_index()  # Python has great order-of-operations (*Brenda)
    sols2 = sols.reset_index(drop=True)

    # Calculate weighted dollar allocations
    w = sols2.mul(weights2, axis=0)
    w['rating'] = sols.reset_index()['rating']
    w['portfolio'] = sols.reset_index()['portfolio']
    w = w.set_index(['portfolio', 'rating'])

    # Group by rating and sum for total allocation
    w_grouped = w.groupby('rating')
    for index, row in w_grouped:
        total_values = row.sum()
        total_values['rating'] = index
        total_values['portfolio'] = 'Total'
        w = pd.concat([w, pd.DataFrame(total_values).T.set_index(['portfolio', 'rating'])])

    return w.reset_index()

''' This function takes in the asset mix and the solved solution up to this point to calculate how much of the total allocation has been allocated in each portfolio. Those weights are used as bounds for the total optimization''' # Old comment vs my interpretation (kept both in case err - for now)
def get_bnds_for_total(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculates the bounds for total optimization based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the upper and lower bounds for optimization for each rating.
    """
    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')
    dollars = Asset_mix['Total']

    # Calculate bounds by dividing total allocation by asset mix
    bounds = total.div(dollars, axis=0)
    bounds = bounds.where(bounds > 0, 0)

    return bounds


def liabilities_table(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a table of liability allocations for each bond rating based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the liability allocations by bond rating.
    """
    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')
    dollars = total.sum(axis=1)

    # Calculate liability allocations by dividing by total dollars
    liabilities = total.div(dollars, axis=0)
    liabilities['rating'] = liabilities.index
    liabilities['portfolio'] = 'Liability'

    return liabilities.reset_index(drop=True)


def surplus_table(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a table of surplus allocations for each bond rating based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the surplus allocations by bond rating.
    """
    npt_weights = Asset_mix['Total']
    npt_sol = solution_df[solution_df['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')

    # Calculate optimized solution using weights
    optimization_sol = npt_sol.mul(npt_weights, axis=0)

    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')

    # Calculate the surplus by subtracting total from optimization
    total = optimization_sol - total
    dollars = total.sum(axis=1)

    # Calculate surplus allocations by dividing by total dollars
    surplus = total.div(dollars, axis=0)
    surplus['rating'] = surplus.index
    surplus['portfolio'] = 'Surplus'

    return surplus.reset_index(drop=True)


from typing import List
from datetime import datetime

def calc_bounds(given_date: datetime, portfolio: str, total: float) -> List[List[float]]:
    """
    Calculates the optimization bounds for cashflow buckets based on historical data for a given portfolio.

    Parameters:
    given_date (datetime): The date to base the historical data on.
    portfolio (str): The portfolio type (such as 'ul' or 'np') for which bounds are being calculated.
    total (float): The total portfolio value to normalize bounds.

    Returns:
    List[List[float]]: A list of bounds for each cashflow bucket, defining the lower and upper limits.
    """
    if portfolio not in ['ul', 'np']:
        return [[0, 1]] * 6  # Default bounds for other portfolios

    year = given_date.strftime('%Y')
    year_folder = given_date.strftime('%Y')
    quarter = ((given_date.month - 1) // 3) + 1
    prev_quarter = quarter - 1
    if prev_quarter == 0:
        prev_quarter = 4
        year = str(given_date.year - 1)

    quarter = str(quarter)
    prev_quarter = str(prev_quarter)

    # Construct the file name based on the quarter and year
    if given_date.year == 2024 and quarter == '1':
        file_name = f"{portfolio} IFE Estimate Q1 2024.xlsx"
    else:
        file_name = f"{portfolio} IFE Estimate Q{quarter} {year}.xlsx"

    path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation", year_folder,
                              'IFE estimates', f'Q{quarter}', file_name)
    try:
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)
    except FileNotFoundError:
        file_name = f"{portfolio} IFE Estimate Q{prev_quarter} {year} to Q{quarter}.xlsx"
        path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation",
                                  year_folder, 'IFE estimates', f'Q{quarter}', file_name)
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    ws = workbook['CF']
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    # Retrieve the present value (PV) for each cashflow bucket
    cf_pvs = df.iloc[1:7, 34].tolist()
    bounds = []

    # Define bounds for each bucket, allowing short positions if PV is negative
    for pv in cf_pvs:
        if pv >= 0:
            bounds.append([0, 6])
        else:
            bounds.append([pv / total, 6])

    return bounds

''' given a df with a multi-index, portfolio and rating, this function will sum all rows with the same rating, and append the sum to a new row with portfolio 'Total' '''
def get_totals_for_rating(df: pd.DataFrame, reset_index: bool = False) -> pd.DataFrame:
    """
    Summarizes the total values for each rating in the given DataFrame by aggregating portfolios.

    Parameters:
    df (pd.DataFrame): A DataFrame with multi-index of 'portfolio' and 'rating'.
    reset_index (bool): Whether to reset the index after aggregating. Default is False.

    Returns:
    pd.DataFrame: A DataFrame with total values for each rating, with an additional row for portfolio 'Total'.
    """
    print(df)

    df_copy = df.copy()
    df_grouped = df_copy.groupby('rating')

    # Sum all rows with the same rating and append the total row
    for index, row in df_grouped:
        total_values = row.sum()
        total_values['rating'] = index
        total_values['portfolio'] = 'Total'
        total_values_df = pd.DataFrame(total_values).T.set_index(['portfolio', 'rating'])
        df_copy = pd.concat([df_copy, total_values_df])

    return df_copy.reset_index() if reset_index else df_copy


# Reads in liabilities from Targets by asset class
# Targets by Asset Class functions:
def public_sensitivities(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves public asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for public asset classes.
    """

    # file_name = "Targets By Asset Class Modified.xlsx"
    file_name = "Targets By Asset Class.xlsx"
    # file_name = "Targets By Asset Class 100%.xlsx"

    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"
    quarter = f"Q{quarter}"

    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    sheet = 'public'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df


def private_sensitivities(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves private asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for private asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    # file_name = "Targets By Asset Class 100%.xlsx"
    # file_name = "Targets By Asset Class Modified.xlsx"

    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)

    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"

    quarter = f"Q{quarter}"


    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    sheet = 'private'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df

def mortgage_sensitivities(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves mortgage asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for mortgage asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    # file_name = "Targets By Asset Class 100%.xlsx"
    # file_name = "Targets By Asset Class Modified.xlsx" # TODO! pull out the targets by asset class for sure

    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)

    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"

    quarter = f"Q{quarter}"


    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)


    sheet = 'mortgage'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df


### HELPERS CODE ###

#class GetData:
#class CreateShocks:


def create_bucketing_table() -> pd.DataFrame:
    """
    Creates a bucketing table with term-to-maturity intervals across 70 buckets.
        - Used in create_weight_tables(ftse_data)
    Returns:
        pd.DataFrame: A DataFrame with term buckets and their respective lower and upper bounds.
    """
    # Create a DataFrame with term buckets ranging from 0.5 to 35 years (70 intervals)
    d = {'Term': list(np.linspace(start=0.5, stop=35, num=70))}
    df = pd.DataFrame(data=d)

    # Calculate the lower and upper bounds for each bucket
    df['Lower_Bound'] = (df['Term'] + df['Term'].shift(
        1)) / 2  # This is equal to calculating the +25, -25 lower_bound and upper_bound, some of which other functions implement.
    df['Upper_Bound'] = df['Lower_Bound'].shift(-1)

    # Adjust the first and last bounds
    df.iloc[0, 1] = 0
    df.iloc[-1, 2] = 100
    # The last bound is 100 to include weights of whole universe (cashflows generated from bonds <= 35 years,
    # however overall weights for KRDs and solution CFs should equate to 100 => hence, bounds of 100 here).
    return df


def create_weight_tables(ftse_data: pd.DataFrame) -> tuple[Dict[str, pd.DataFrame], pd.DataFrame]:
    """
    Creates weight tables for each bond rating based on FTSE Universe (subindex) percentages.

    Usage:
    Used by make_KRD_tables to aggregate bonds into 6 buckets.

    Parameters:
    ftse_data (pd.DataFrame): A DataFrame containing bond information from the FTSE universe.

    Returns:
    weight_dict (Dict[str, pd.DataFrame]): A dictionary of weight tables for each bond rating.
    total_universe_weights (pd.DataFrame): A DataFrame with total market weights for each rating and term bucket.
    """
    buckets = [1, 5.75, 10.75, 15.75, 20.75, 27.75, 35.25]  # Predefined term buckets
    weight_dict = {}

    total_universe_weights = pd.DataFrame(
        index=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        columns=list(range(1, 7)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        column_to_look_in = "RatingBucket" if rating != 'Corporate' else "Sector"

        # Create bucketing table to weigh 70 buckets into 6 buckets.
        df = create_bucketing_table()

        # Sum market weights within each bucket
        for x in range(6):
            lower_b = buckets[x]
            upper_b = buckets[x + 1]
            column_name = f"{lower_b} - {upper_b}"

            # Calculate total market weight for the given rating and term bucket
            df[column_name] = df.apply(lambda row: ftse_data.loc[
                (ftse_data[column_to_look_in] == rating) &
                (ftse_data['TermPt'] < upper_b) &  # if between lower and upper bounds && between the Lower and Upper bounds by create bucketing table
                (ftse_data['TermPt'] >= lower_b) &
                (ftse_data['TermPt'] < row['Upper_Bound']) &
                (ftse_data['TermPt'] > row['Lower_Bound'] - 0.0001)
                ]['marketweight_noREITs'].sum(), axis=1)

            total_universe_weights.loc[rating, x + 1] = df[column_name].sum()

            # Normalize by the sum of market weights
            df[column_name] = df[column_name] / df[column_name].sum()

        weight_dict[rating] = df # NaNs exist here, potentially.

    return weight_dict, total_universe_weights


# class CreateShocks:
def create_general_shock_table() -> pd.DataFrame:
    """
    Creates a reusable shock table to calculate shocks for each security type.

    Creates n by m matrix, where
        n = 70 buckets (for semi-annual bond yields across 35 years)
        m = 11 columns of shock intervals

    Each row sums up to 1 basis point.

    Returns:
    pd.DataFrame: A DataFrame containing the shock values for different term buckets.
    """
    shock_size = 0.0001  # We define 1 basis point here as 0.0001.

    buckets = [0, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30, 100]  # Shock interval years.

    terms = np.linspace(start=0.5, stop=35, num=70)  # Buckets for bond terms.

    # Create a DataFrame for shocks with 70 term intervals
    shocks = pd.DataFrame(index=terms, columns=buckets, dtype=float)
    shocks[0] = terms  # Initialize the first column with term points

    # Calculate shocks for each bucket
    for i in range(1, 11):
        # Define masks for terms within the left and right ranges
        left_mask = (buckets[i - 1] <= terms) & (terms <= buckets[i])
        right_mask = (buckets[i] < terms) & (terms <= buckets[i + 1])  # Notice the strict inequality here
        # Calculate left and right shocks
        left_shock = (terms - buckets[i - 1]) / (buckets[i] - buckets[i - 1]) * shock_size
        right_shock = (1 - (terms - buckets[i]) / (buckets[i + 1] - buckets[i])) * shock_size
        # Assign shocks, ensuring no overlap
        shocks[buckets[i]] = np.where(left_mask, left_shock, 0)
        shocks[buckets[i]] = np.where(right_mask, right_shock, shocks[buckets[i]])

    # Manual overwrite as per original function change to ensure the sum of 1 basis point in early and end bucket terms.
    shocks.iloc[0, 1] = shock_size
    shocks.iloc[60:70, 10] = shock_size

    # Drop the last column (bucket 100)
    shocks = shocks.drop(columns=100)

    # Time of running this function is, on average, 0.00000 seconds.
    return shocks


def create_shock_tables(semi_annual_curves, GivenDate: datetime, debug=False) -> dict[str, Any]:
    """
    Applies the shocks to the bond curves for each rating and store results in shocks_dict.
    """

    # makes a dictionary containing tables for up shocks and down shocks for each rating
    shocks_dict = {}
    up_shocks = create_general_shock_table()  # creates a df with col named '0', '1', etc


    cur_date = GivenDate.strftime('%Y%m%d')

    folder_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), 'Benchmarking', 'Test', 'benchmarking_outputs',
                               'Brenda', 'shocks_table')

    excel_filename = 'shocks_table'
    file_path = os.path.join(folder_path, f'{excel_filename}_{cur_date}.xlsx')

    if not os.path.exists(folder_path):
        os.mkdir(folder_path)

    if not os.path.exists(file_path):
        with pd.ExcelWriter(file_path) as writer:
            up_shocks.to_excel(writer, sheet_name='general_shocks')
    else:
        print('shocks file for this quarter already exists - cant make a file with the same name')



    down_shocks = create_general_shock_table()
    down_shocks = -down_shocks  # can decouple into classes
    down_shocks[0] = -down_shocks[0]
    curves_mod = semi_annual_curves

    cur_date = GivenDate.strftime('%Y%m%d')  # givendate to str - consider doing it here or as fn

    CURR_DEBUGGING_PATH = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), 'Benchmarking', 'Test',
                                       'benchmarking_outputs', 'Brenda', cur_date, 'Debugging_Steps')
    os.makedirs(CURR_DEBUGGING_PATH, exist_ok=True)
    file_utils.write_results_to_excel(curves_mod, CURR_DEBUGGING_PATH, cur_date, 'interpolated_bond_curves')


    # Apply up and down shocks to bond curves
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for direction in ['Up', 'Down']:
            table_name = rating + ' - ' + direction
            if direction == 'Up':
                df = up_shocks
            else:
                df = down_shocks

            # Add shocks to bond curves
            df = df.add(curves_mod[rating], axis=0)
            df[0] = curves_mod[rating]

            # Apply power function to simulate the bond curve transformation after shocks
            df += 1
            df['Powers'] = df.index
            df = df.pow(df['Powers'], axis=0)
            df = 1 / df
            df = df.drop('Powers', axis=1)

            shocks_dict[table_name] = df

    return shocks_dict


# """
# Interpolating for half-years - side-eff 1
# """

# TODO: temp function, can split up functionality
def create_semi_annual_bond_curves(curves) -> pd.DataFrame:  # curves from ftse curves
    # Interpolating bond curves for half-year intervals (linear interpolation; take average of up-down years)
    curves_mod = pd.DataFrame(
        columns=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        index=list(np.linspace(start=0.5, stop=35, num=70)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for i in range(1, 35):
            curves_mod.loc[i, rating] = curves.loc[i, rating]
        for yr in np.linspace(start=1.5, stop=34.5, num=34):
            curves_mod.loc[yr, rating] = (curves.loc[yr + .5, rating] + curves.loc[yr - .5, rating]) / 2

    curves_mod.loc[0.5] = curves_mod.loc[1]
    curves_mod.fillna(method='ffill', inplace=True)

    return curves_mod



# Does NOT modify bond curves parameter; assigns to new value from reference data by dereferencing value



# (*begin) Takes each year and looks at rating and FTSE universe (half-year would be from .25 to .75; up quarter year and down quarter year for half year, and so on for every year

##### version 1-OLD begins #####

# Function to calculate the average coupon rate for a specific bond rating and year
# It uses the FTSE data to filter bonds based on the given rating and term (maturity year).
# The average coupon is weighted by the notional weight of the bond, excluding REITs.
# The 'price' is MVAI (market weighted price after interest), and we divide it out so removed the market weighting to retain the
# Notional weighting

# TODO! This actually buckets it for the cashflows
#  this function is directly related to the create_cashflows_70 function and does the bucketing for it
def calc_avg_coupon(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the average coupon rate for a specific bond rating and year, weighted by the notional weight of the bond.

    Parameters:
    year (float): The specific year (maturity) to calculate the coupon for. Called 'Bucket", a bucket of 0.5 increments from 0.5 ttm to 35 ttm
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The average coupon rate for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) &  # TODO! essentially, the cashflows 70 uses an appropriate bucketing method in calc_pv and calc_avg_coupon and the weights (6) uses create_bucketing_table to determine the 6th bucket weightings from the FTSE universe, which uses an upper bound of 100 - this mainly affects provincial bond sensitivity weightings, as, the provincial bonds comprise more of the ttms > 35.25. notice that this is bucket 0 in ftse universe, already determined - so this function could simplify it further. Moreover, the 70 cashflows use a different weighting system, and, I presume are less sensitive in 6 weightings than the KRD sensitivities - should have a more formalized (or ocnsistent) system of weighting the exact same way, imo
                       (ftse_data['marketweight_noREITs'] > 0)]   # TODO! NOTE: this uses a DIFFERENT bucketing system than create_bucketing_tables() which is used for the bounds of calculating the 6 weights from 70 tables. Lol, this is funny

    # If no bonds match the criteria, return a coupon rate of 0
    if df.empty:
        return 0

    # Otherwise, calculate the weighted average coupon rate, dividing by 2 for semi-annual coupon payments. As follows:
        # 1. Multiply the market weight by the coupon rate and divide by the market value-adjusted interest (mvai).
        # 2. Divide the sum of these weighted values by the sum of market weights/mvai.
    avg_coupon = ((df['marketweight_noREITs'] * df['annualcouponrate'] / df['mvai']).sum() / # Change in code to use the price (SAME as EXCEL) ****
                  (df['marketweight_noREITs'] / df['mvai']).sum()) / 2  # Divide by 2 to account for semi-annual coupons
    # second SUMPROD is notional weighting
    # Return the calculated average coupon rate for the given rating and year (average was 0 if no matching bonds were found from FTSE bond databank)

    return avg_coupon

##### version 1-OLD ends #####

""" Notional Weighting """


# TODO: Cashflows * interpolated (unshocked) curves


# Function to calculate the present value (PV) of bonds for a specific rating and year
# It uses the FTSE data to filter bonds based on the rating and term and then calculates the PV.
def calc_pv(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the present value (PV) of bonds for a specific bond rating and year.

    Parameters:
    year (float): The specific year (maturity) to calculate the PV for.
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The present value (PV) for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) &
                       (ftse_data['marketweight_noREITs'] > 0)]

    # If no bonds match the criteria, return a PV of 0
    if df.empty:
        return 0

    # Otherwise, calculate the present value (PV) by summing up the product of the market weight and the bond's 
    # market value-adjusted interest (mvai), then dividing by the sum of the market weights.
    
    # I.e., Calculate the present value (PV) as the weighted sum of market value-adjusted interest (mvai)
    # TODO: change to excel!
    pv = (df['marketweight_noREITs'] * df['mvai']).sum() / df['marketweight_noREITs'].sum()

    # Return the calculated present value for the given rating and year
    return pv


def create_cf_tables(ftse_data):
    # uses the average coupon rate to calculate annual cashflows for each rating type
    cf_dict = {}
    years = list(np.linspace(start=0.5, stop=35, num=70))
    buckets = list(np.linspace(start=0.5, stop=35, num=70))
    df = pd.DataFrame(columns=years, index=buckets)
    df.insert(0, 'Bucket', buckets)
    df.insert(1, 'Principal', 100)

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:

        df = pd.DataFrame(columns=years,index=buckets)
        df.insert(0, 'Bucket', buckets)
        df.insert(1, 'Principal', 100)

        df['PV'] = df.apply(lambda row: calc_pv(row['Bucket'], rating, ftse_data), axis=1)
        df['Coupon'] = df.apply(lambda row: calc_avg_coupon(row['Bucket'], rating, ftse_data), axis=1)

        coupons = df.pop(df.columns[-1])
        df.insert(2, 'Coupon', coupons)

        for col in np.linspace(start=0.5, stop=35, num=70):
            df[col] = df.apply(lambda row: row['Coupon'] if row['Bucket'] > col else ((row['Coupon'] + row['Principal']) if row['Bucket'] == col else 0), axis=1)

        cf_dict[rating] = df.iloc[:, :73]
        cf_dict[rating + 'PV'] = df.iloc[:, 73]

    return cf_dict

# Input: ftse_data - a DataFrame containing bond information.
# Output: cf_dict - a dictionary of cashflow tables and their respective present values for each bond rating.

def create_sensitivity_tables(cashflows: Dict[str, pd.DataFrame], shocks: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Calculates cashflow sensitivities based on shocks applied to bond curves.

    Parameters:
    cashflows (Dict[str, pd.DataFrame]): A dictionary containing cashflow tables for each bond rating.
    shocks (Dict[str, pd.DataFrame]): A dictionary containing shock tables for each bond rating.

    Returns:
    Dict[str, pd.DataFrame]: A dictionary of sensitivity tables for each bond rating.
    """
    sensitivities_dict = {}  # Dictionary to store sensitivity tables
    buckets_krd = [0, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30]  # KRD buckets

    # Iterate through each bond rating type to calculate sensitivities
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        # Retrieve cashflows and shocks for the current rating
        cfs = cashflows[rating]  # Retrieve cashflows for the current rating
        ups = shocks[rating + ' - Up']  # Retrieve up shock table for the current rating
        downs = shocks[rating + ' - Down']  # Retrieve down shock table for the current rating

        """
        ## sumproduct for each, changed to get the sensitivities
        # cahsflows for the square (70*70) table, and it fits into the 10*70 sensitivities that it matches up to - sum to each one, cahsflow*shocks.
        """
        # Create empty DataFrames to store sensitivity data for up and down shocks
        df_up = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_up.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        df_down = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_down.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        # Calculate sensitivities by summing the product of cashflows and shocks
        for x in range(1, 11):
            for i in range(70):
                df_up.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * ups.iloc[:, x])  # Multiply cashflows by up shocks
                df_down.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * downs.iloc[:, x])  # Multiply cashflows by down shocks

        # Calculate the average sensitivity (difference between down and up shocks divided by 2)
        average_sensitivity = (df_down - df_up) / 2 * 10000

        # Add bucket information and transpose for better readability
        average_sensitivity['Bucket'] = list(np.linspace(start=.5, stop=35.5, num=71))
        average_sensitivity = average_sensitivity.transpose()
        average_sensitivity = average_sensitivity.drop(70, axis=1)
        average_sensitivity = average_sensitivity.iloc[1:]

        # Insert bucket names for KRD
        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        # TODO: NEW CODE
        avg_sensitivity = average_sensitivity
        # End

        for x in range(10):
            for i in range(70):
                # TODO: for the PV stuff, old code:
                """
                # Safe division, handling division by zero and inf
                numerator = average_sensitivity.iloc[x, i + 1]
                denominator = cashflows[rating + 'PV'].iloc[i]

                # Use np.divide with where clause to avoid division by zero and handle inf
                average_sensitivity.iloc[x, i + 1] = np.divide(numerator, denominator, out=np.zeros_like(numerator), where=denominator != 0)  # Gets the dollar-weighted amounts
                """
                # TODO: NEW CODE
                pv = np.sum(cashflows[rating].iloc[i, 3:] * ups.iloc[:, 0]) # it selects the row, nice (row, which are a bucket) - and ups.iloc[:,0] holds the PV values; of discounted curves
                average_sensitivity.iloc[x, i + 1] = avg_sensitivity.iloc[x, i + 1] / pv
                # End
        # Store the calculated sensitivity table for the rating
        sensitivities_dict[rating] = average_sensitivity


    return sensitivities_dict




def make_krd_table(weights: Dict[str, pd.DataFrame], sensitivities: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Creates the final Key Rate Duration (KRD) table by combining market weights and cashflow sensitivities for each bond rating.

    Parameters:
    weights (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing market weights for each bond rating and maturity bucket.
    sensitivities (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing sensitivities for each bond rating and maturity bucket.

    Returns:
    pd.DataFrame: A combined KRD table for all bond ratings and term buckets.
    """
    KRDs = {}
    cols = ['rating', 'term', 'bucket1', 'bucket2', 'bucket3', 'bucket4', 'bucket5', 'bucket6']
    buckets = [1, 2, 3, 5, 7, 10, 15, 20, 25, 30]

    # Iterate over each bond rating to calculate KRD values
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        df = pd.DataFrame(columns=cols, index=range(10))
        df['term'] = buckets  # Assign bucket terms to the DataFrame
        df['rating'] = rating  # Set bond rating

        # Calculate KRD by multiplying sensitivities with market weights for each bucket
        for x in range(2, 8):
            df.iloc[:, x] = df.apply(lambda row: (
                sensitivities[rating].loc[sensitivities[rating]['Bucket'] == row['term']].iloc[:, 1:].values[0] *
                weights[rating].iloc[:, (x + 1)]
            ).sum(), axis=1)

        KRDs[rating] = df  # Store KRD DataFrame in the dictionary

    # Concatenate all rating-specific KRD DataFrames into one final DataFrame
    final_krd_df = pd.concat([KRDs['Federal'], KRDs['Provincial'], KRDs['CorporateAAA_AA'], 
                              KRDs['CorporateA'], KRDs['CorporateBBB'], KRDs['Corporate']], ignore_index=True)

    final_krd_df.fillna(0, inplace=True)  # Replace NaN values with 0
    return final_krd_df



