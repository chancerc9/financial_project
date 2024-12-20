"""
    This provided code is a financial script that processes bond-related data,
    calculates key rate durations (KRDs), and runs an optimization script for asset-liability
    matching - see specific function comments for most up-to-date and accurate descriptions - and expected returns
    optimization (for totals). This produces the percent allocations for investing in a model portfolio - used for
    investment benchmarking purposes.
"""
# Standard library imports
import datetime as dt
import os
import sys

# Third-party imports
import numpy as np
import pandas as pd
from scipy import interpolate
import openpyxl
from typing import List

# Local application-specific imports
from equitable.infrastructure import sysenv
from equitable.utils import processtools as misc
from scipy.optimize import minimize

# Adding system path for custom imports
sys.path.append(sysenv.get("ALM_DIR"))

# Required custom modules
import calculations as helpers
import file_utils as file_utils
import datahandler
import helpers as function_helpers

# Configure pandas display settings
pd.set_option('display.width', 150)


# --- Start of Model Portfolio code: ---

"""
1. Create Asset KRDs

Function: create_AssetKRDs()

Create Asset KRDs to be used in optimization procedure below.

Mutates: Nothing

Side effects: 
    - calls make_krd_table() to create 6 bucket KRDs from 70 bucket KRDs (sensitivities), using weights to scale and allocate them.
    - this function does *not* mutate/alter/change the original bond_curves dataframe reference, nor any further changes to replicates.
"""
def create_AssetKRDs(const_bond_curves: pd.DataFrame, ftse_data: pd.DataFrame,
                       GivenDate: pd.Timestamp, debug=False) -> pd.DataFrame: # DEBUGGING_DIRECTORY: str=None,
    """
    Creates the Key Rate Duration (KRD) table for assets on a given date.
    (Main method to create the KRD table for assets.)

    Parameters:
    GivenDate (pd.Timestamp): The date for which to calculate the KRD table.

    What it does:
    Creates and aggregates the KRD profiles (i.e., sensitivities) and weighted-average it into 6 buckets.

    Elaboration:
    Calculates the KRD profiles (i.e., sensitivities) and calls make_krd_table(sensitivities) to perform a weighted-averages
    for the sensitivities into 6 buckets. Final df of KRD profiles for 6 buckets is used for optimizer and produced for KRD
    profiles solutions results.

    Returns:
    pd.DataFrame: A DataFrame containing weighted sensitivities for each bond rating. For 6 buckets; used for optimizer.
    """

    # Create weight tables, cashflow tables, shock tables, and sensitivity tables; makes a weight table for the 6 buckets (to 6 buckets, from the 70 buckets cfs)
    weights, totals = helpers.create_weight_tables(ftse_data)                   # Makes a weight table that holds weights for each bond rating and bucket
    cf_tables = helpers.create_cf_tables(ftse_data)                             # Makes a 30-35 year average semi-annual cashflow table for each bond rating and bucket, with principal 100.
    shock_tables = helpers.create_shock_tables(const_bond_curves, GivenDate)    # Makes 30 year up and down shock tables for each bond rating and bucket.

    # Creates cashflow sensitivities for assets: sensitivities are in 70 ttm buckets * 10 KRD shock intervals (terms)
    sensitivities = helpers.create_sensitivity_tables(cf_tables, shock_tables)  # Uses shocks and cashflows to make 30 year sensitivity tables for each bond rating and bucket
                                                                                # sensitivities are in 70 ttm buckets * 10 KRD shock intervals (terms)
                                                                                # Use this and weights to make final KRD tables (same thing but 6 buckets)

    if debug:
        # Ensure debugging directory exists for storing the results for debugging purposes:
        cur_date = GivenDate.strftime('%Y%m%d')
        general_debugging_path: str = function_helpers.build_and_ensure_directory('Benchmarking',
                                                                                  'code_benchmarking_outputs',
                                                                                  cur_date, 'debugging_steps')

        # Output directory for asset sensitivities of aggregate and granular buckets:
        asset_sensitivities_path: str = function_helpers.build_and_ensure_directory(general_debugging_path,'Asset_KRDs')

        # Granular sensitivities
        path: str = function_helpers.build_and_ensure_directory(asset_sensitivities_path, 'granular_sensitivities')

        # Save asset sensitivity tables as Excel files for each rating for these 70 buckets:
        for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
            file_path = os.path.join(asset_sensitivities_path, f'{rating}_70_bucket_asset_sensitivities_{cur_date}.xlsx')
            if not os.path.exists(file_path):
                with pd.ExcelWriter(file_path) as writer:
                    sensitivities[rating].to_excel(writer)

    # --- Calculations ---

    # Weighs the sensitivities table (70 buckets) into overall KRD table (6 buckets) with FTSE universe market weights
    df = helpers.make_krd_table(weights, sensitivities)
    df['rating'] = df['rating'].replace({
        'CorporateBBB': 'corporateBBB',
        'CorporateA': 'corporateA',
        'CorporateAAA_AA': 'corporateAAA_AA'
    })

    print("Created the Asset KRDs")

    """
    Method for debugging:
    """
    if debug:

        file_utils.write_results_to_excel(const_bond_curves, general_debugging_path, cur_date,
                                          'ftse_bond_curves_semiannual')

        # Creates weight tables for each bond rating based on subindex percentages (for bonds):
        file_utils.write_results_to_excel(weights, general_debugging_path, cur_date,
                                          'bond_weights_per_rating_for_6_buckets')  # Weighting to make 70 buckets into 6 buckets
        file_utils.write_results_to_excel(totals, general_debugging_path, cur_date,
                                          'total_universe_weights')

        # Shocked curves table:
        file_utils.write_results_to_excel(shock_tables, general_debugging_path, cur_date, 'shocked_bond_curves')

        # KRD table (final KRDs):
        file_utils.write_results_to_excel_one_sheet(df, asset_sensitivities_path, cur_date,'final_krd_table')

        # cf tables based on ftse data:
        file_utils.write_results_to_excel(cf_tables, general_debugging_path, cur_date, 'cf_tables_ftse_data')

    """
    End of method for debugging
    """

    return df


"""
Function: reading_asset_mix()
    Helper that create_asset_KRDs(ftse_data, GivenDate) relies on.

Mutates: Nothing

Information:
    To determine the asset types per segment, multiply the segment balance by the asset mix percentage that is unique to each
    segment. This percentage matrix is called the "Asset Mix.xlsx" file, where the segment balances and total asset balance is 
    called the "SBS Totals.xlsx" file.
"""
# Should be called: create asset_mix
def reading_asset_mix(Given_date: pd.Timestamp, sheet_version: int = 1) -> tuple[
    pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Creates the asset mix balance for bond ratings (private, public, and mortgage portfolios) by multiplying
    segment balances and asset mix matrix (percentage of each rating in each segment).

    Parameters:
    Given_date (pd.Timestamp): The date for which the quarter's asset mix is retrieved and calculated.
    sheet_version (int, optional): Determines if totals or segments are calculated (1 for segments, 0 for totals (optimization) - defaults to 1).

    Returns: a Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame] of three DataFrames
        - df_public: Public assets.
        - df_private: Private assets.
        - df_mortgages: Mortgages.
    """

    # --- Get Source Data ---

    # Assumptions:
        # 'SBS Totals.xlsx': Aggregated segment and total balances, for company assets.
        # 'Asset Mix.xlsx': Percentage matrix contains the percents for ratings (rows) per segment (columns), for company assets.

    # Obtain balance sheet aggregated totals or aggregated segment totals for all equitable assets:
    totals = datahandler.get_BSTotals(Given_date, sheet_version) # sheet_version = 1 for segments, sheet_version = 0 for totals

    # Obtain the percentage matrix for company assets (of bond ratings, per segment, for equitable):
    weights = datahandler.get_percents(Given_date)
    weights = weights[['ACCUM', 'PAYOUT', 'UNIVERSAL', 'NONPAR', 'GROUP', 'PARCSM', 'Total', 'Surplus', 'SEGFUNDS']]


    # --- Pre-process Source Data ---

    # Remove columns where every entry of the column is NaN:
    weights = weights.dropna(axis=1, how='all')


    # --- Calculate Results ---

    # Multiply the balance by asset percents of segments to get asset mix:
    df = weights.multiply(pd.Series(totals))
    df.index.name = None

    # Renaming:
    df.rename(columns={'ACCUM': 'Accum', 'PAYOUT': 'Payout', 'GROUP': 'group', 'UNIVERSAL': 'ul', 'NONPAR': 'np'},
              inplace=True)


    # --- Split results into private, mortgage, and public tables, and rename for convention ---

    # Split into public, private, and mortgage tables, defining them (as per ALM team):

    # Public bonds:
    df_public = df.iloc[:5].copy()
    df_public.rename({'CorpAAA_AA': 'CorporateAAA_AA', 'CorpA': 'CorporateA', 'CorpBBB': 'CorporateBBB'},
                     inplace=True)

    # Privates are modelled with CorpA and CorpAAA bonds (as private assets DNE in FTSE Universe):
    df_private = df.iloc[5:11].drop(columns=['SEGFUNDS']) # .drop() explicitly creates a new df, hence, avoiding
                                                          # ambiguity or warnings from Pandas.
    df_private.rename({'PrivateAA': 'CorporateAAA_AA', 'PrivateA': 'CorporateA', 'PrivateBBB': 'CorporateBBB',
                       'MortgagesInsured': 'Federal'}, inplace=True)

    # Mortgages are modelled from privates:
    # Define mortgages from private_df rows:
    df_mortgages = df_private.loc[['Federal', 'MortgagesConv']].copy()
    df_mortgages.rename({'MortgagesConv': 'CorporateBBB'}, inplace=True)

    # Final dropping of rows to define df_private:
    df_private.drop(['PrivateBB_B', 'MortgagesConv', 'Federal'], inplace=True)

    return df_public, df_private, df_mortgages



"""
2. 

Purpose:
Brings in liability sensitivities (from the ALM, liabilities team) and performs an asset-liability matching and an
optimization procedure to produce asset allocation percentages for a portfolio model used in benchmarking.

Functions: 
optimization() is a wrapper for bringing the separate optimization procedures together (can merge if function weights or 
                calculation changes to reflect excel)
optimization_worker() performs segment level and total portfolio level (with segment weights of 1) optimization on 
                asset and liability sensitivities to create a portfolio model (outputs as solutions.xlsx, 
                Model Portfolio; used for benchmarking in Custom_benchmarks.xlsx, and to generate the model portfolio 
                cashflows from solutions).

These are functions that *may* mutate things. 

Overview: 
    Performs optimization calculation, including asset-liability sensitivity matching
    and optimizing expected returns for totals (using segment weights of 1, for the totals portfolio).  
"""
def optimization(AssetKRDs: pd.DataFrame, given_date: dt, LOGFILE, asset_type='public'):
    """
    Purpose:
    optimization(AssetKRDs, given_date, LOGFILE, asset_type) is a wrapper function for optimization_worker(),
        which performs an asset-liability hedging and optimization:
        optimizes (matches) assets to liabilities for segments and optimizes expected returns
        for totals.

    Misc:
        Passes down LOGFILE for logging purposes.
        optimization_worker function creates a copy of KRDs to prevent the propagation of changes.

    AssetKRDs: 6 bucket KRDs for optimization process.
    LOGFILE: for logging purposes.
    asset_type: specify type of asset to optimize.
    """

    # Obtain optimized solution dfs (does not modify the reference of AssetKRDs):

    # Segments:
    sol_df_seg = optimization_worker(AssetKRDs, given_date, LOGFILE, asset_type, sheet_version=1)  # segments = 1
    # Totals:
    sol_df_tot = optimization_worker(AssetKRDs, given_date, LOGFILE, asset_type, sheet_version=0)  # totals = 0

    def overwrite_total_rows(sol_df_seg, sol_df_tot):
        """
        Overwrite 'Total' portfolio rows in sol_df_seg with rows from sol_df_tot.

        Args:
        sol_df_seg: DataFrame containing segment results (public, private, mortgage).
        sol_df_tot: DataFrame containing total portfolio results.

        Returns:
        sol_df_seg: Updated DataFrame with 'Total' portfolio rows replaced by sol_df_tot rows.
        """

        # Step 1: Filter out the 'Total' rows from both sol_df_seg and sol_df_tot
        total_rows_tot = sol_df_tot[sol_df_tot['portfolio'] == 'Total']
        non_total_rows_seg = sol_df_seg[sol_df_seg['portfolio'] != 'Total']

        # Step 2: Concatenate non-'Total' rows from sol_df_seg with 'Total' rows from sol_df_tot
        updated_sol_df_seg = pd.concat([total_rows_tot, non_total_rows_seg], ignore_index=True)

        return updated_sol_df_seg

    sol_df = overwrite_total_rows(sol_df_seg, sol_df_tot)


    # Renaming conventions can place here: #TODO! Test
    #benchmarking_solution = solution_df.copy()
    #FTSE_Universe_data = FTSE_Universe_data.copy()
    #ftse_data = IndexTable.copy()
    #bond_curves = bond_curves.copy()

    # Rename columns and portfolios in benchmarking_solution_df
    ratings_map = {
        'Federal': 'Federal',
        'Provincial': 'Provincial',
        'corporateAAA_AA': 'CorporateAAA_AA',
        'corporateA': 'CorporateA',
        'corporateBBB': 'CorporateBBB'
    }
    portfolio_map = {
        'Total': 'TOTAL',
        'np': 'NONPAR',
        'group': 'GROUP',
        'Accum': 'ACCUM',
        'Payout': 'PAYOUT',
        'ul': 'UNIVERSAL',
        'Surplus': 'SURPLUS'
    }
    # Rename columns mapping
    rename_buckets_map = {5: 6, 4: 5, 3: 4, 2: 3, 1: 2, 0: 1}

    sol_df.rename(columns=rename_buckets_map, inplace=True)
    sol_df['portfolio'] = sol_df['portfolio'].replace(portfolio_map)  # In place.
    # sol_df['rating'] = sol_df['rating'].str.replace(r'^([a-zA-Z])',lambda m: m.group(1).upper(),regex=True)

    sol_df['rating'] = sol_df['rating'].replace(ratings_map)

    return sol_df


def optimization_worker(AssetKRDsTable: pd.DataFrame, given_date: dt, LOGFILE, asset_type='public', sheet_version=1):

    """
    This optimization function performs asset-liability hedging and optimization procedure:
        optimizes asset sensitivities to liability sensitivities for segments and optimizes expected returns
        for totals.

    With the current function utility, use 1 for segments and 0 for totals (since total weights of segments should be 1 for calculations).

    Simplified the optimization code. Future suggestions include to include appropriate bounds instead of Python clips to bounds.

    Only one asset class is optimized at a time, out of 'public, 'private', 'mortgage'
    """

    # Important: Creates a copy of Asset KRDs table.
    KRDs = AssetKRDsTable.copy()

    # Get assets mix:
    df_public, df_private, df_mortgages = reading_asset_mix(given_date, sheet_version)

    ''' Setting Asset_mix to the correct table based on the given asset class '''
    if asset_type == 'private':
        Asset_mix = df_private.rename(
            {'CorporateAAA_AA': 'corporateAAA_AA', 'CorporateA': 'corporateA', 'CorporateBBB': 'corporateBBB'})

    elif asset_type == 'mortgage':
        Asset_mix = df_mortgages.rename(
            {'CorporateAAA_AA': 'corporateAAA_AA', 'CorporateA': 'corporateA', 'CorporateBBB': 'corporateBBB'})

    else:
        Asset_mix = df_public.rename(
            {'CorporateAAA_AA': 'corporateAAA_AA', 'CorporateA': 'corporateA', 'CorporateBBB': 'corporateBBB'})

    # Get target liability sensitivities from excel:
    ''' Setting the sensitivities to be used as targets for the optimizer, for the chosen asset class'''
    net_sensitivity = datahandler.get_liability_sensitivities(given_date, asset_type)

    if asset_type == 'private':
        # net_sensitivity = helpers.private_sensitivities(given_date)
        total_sensitivity = net_sensitivity.copy()

    elif asset_type == 'mortgage':
        # net_sensitivity = helpers.mortgage_sensitivities(given_date)
        total_sensitivity = net_sensitivity.copy()

    #else:
        ''' For the public optimization, we subtract the private and mortgage target sensitivities from the public target and optimize for the net sensitivity '''
        #net_sensitivity = helpers.public_sensitivities(given_date)

    ''' For the sensitivity targets for the public totals, we subtract the public and mortgage components of all ratings
    we sum the public sensitivities for all 5 portfolios, then subtract the sum of privates for all portfolios, including ParCSM and Surplus'''
    #if asset_type == 'public':
        #net_sensitivity = helpers.public_sensitivities(given_date)

    #elif asset_type == 'private':
        #net_sensitivity = helpers.private_sensitivities(given_date)
    #elif asset_type == 'mortgage':
        #net_sensitivity = helpers.mortgage_sensitivities(given_date)

    solution_df = pd.DataFrame()
    solved_dollar_sensitivities = pd.DataFrame()

    # df for targets (test output)
    krd_targets = pd.DataFrame()

    ''' This df is a table of expected returns taken from the Parallel_tilt_curve_history'''
    expected_return = get_expected_returns()

    ''' start the optimization process for each portfolio'''
    for portfolio in ['ul', 'Payout', 'Accum', 'group', 'np', 'Total']:

        ''' first get the target sensitivities from the df generated above for the current portfolio '''
        df_portfolio = net_sensitivity[net_sensitivity['portfolio'] == portfolio]

        ''' Next, go through each rating class to optimize for each. Calculate provinical last because the target 
        for total-provincial is calculated using the solution for the other ratings '''
        for rating in ['corporateBBB', 'corporateA', 'Federal', 'corporateAAA_AA', 'Provincial']:

            ''' The mortgage portfolios only include Federal and CorporateBBB, and the private doesn't include Fedearl or Provincial. Those cases are excluded from the optimization'''
            if ((asset_type == 'mortgage') & (
                    (rating == 'corporateAAA_AA') or (rating == 'corporateA') or (rating == 'Provincial'))) or (
                    (asset_type == 'private') & ((rating == 'Federal') or (rating == 'Provincial'))):
                continue

            ''' The following cases do not run through the optimizer '''
            if (asset_type == 'public'):
                if ((portfolio == 'np') or (portfolio == 'ul') or (portfolio == 'Payout')):
                    ''' CorporateBBB for Nonpar, Universal and Payout is not optimized. Buckets 3-6 are distributed according to the pre-determined weights to reduce concentration after buckets 1 and 2 are made.
                    CorporateA bonds are also not optimized for Nonpar and Universal - minimum amount is placed in buckets 1 and 2 and remaining is placed in bucket 6'''
                    if (rating == 'corporateBBB') or ((rating == 'corporateA') & (portfolio != 'Payout')):
                        ''' First get the amount to be placed in the first 2 buckets'''
                        bnds = calc_bounds(given_date, portfolio, sum(
                            Asset_mix[portfolio]))
                        new_row_df = pd.DataFrame(0.0, index=np.arange(1), columns=[0, 1, 2, 3, 4, 5])
                        new_row_df.iloc[0, 0] = bnds[0][0]
                        new_row_df.iloc[0, 1] = bnds[1][0]
                        if (rating == 'corporateBBB'):
                            ''' For corporateBBB, follow the weight distribution'''
                            new_row_df.iloc[0, 2:] = [val * (1 - new_row_df.iloc[0, 0:2].sum()) for val in
                                                      [0.1549, 0.2566, 0.4351, 0.1534]]
                        elif (rating == 'corporateA'):
                            ''' For corporateA, place remaining weight in bucket 6'''
                            new_row_df.iloc[0, 5] = 1 - new_row_df.iloc[0, 0:2].sum()
                        ''' Then add the row to the df'''
                        new_row_df['portfolio'] = portfolio
                        new_row_df['rating'] = rating
                        solution_df = pd.concat([new_row_df, solution_df], ignore_index=True)
                        continue

                elif (portfolio == 'Total'):
                    ''' CorporateAAA_AA and Federal in the Total portfolio are not optimized, the remaining investment allocation goes to bucket 6 for Federal, and bucket 1 for CorporateAAA_AA '''
                    if ((rating == 'corporateAAA_AA')):
                        ''' First we get the starting point already calculated by the optimizer for the 5 portfolios '''
                        total_bnds = helpers.get_bnds_for_total(Asset_mix,
                                                                solution_df)

                        new_row_df = total_bnds.loc[[rating]].reset_index(drop=True)
                        new_row_df.iloc[0, 0] = 1 - sum(new_row_df.iloc[0, 1:])

                        new_row_df['portfolio'] = portfolio
                        new_row_df['rating'] = rating
                        solution_df = pd.concat([new_row_df, solution_df], ignore_index=True)
                        continue

            ''' First grab the KRDs of the assets of the corresponding rating '''
            krd = KRDs[KRDs['rating'] == rating]
            krd = krd.reset_index().drop(krd.columns[[0, 1]], axis=1).drop('index', axis=1).to_numpy()

            ''' The get the allocated investment amount for the current bond rating and portfolio'''
            investment_val = Asset_mix[portfolio].loc[rating] / 10000
            ''' If zero, add a blank row to the solution_df '''
            if investment_val == 0:
                new_row_df = pd.DataFrame(0, index=np.arange(1), columns=[0, 1, 2, 3, 4, 5])
                new_row_df['portfolio'] = portfolio
                new_row_df['rating'] = rating
                solution_df = pd.concat([new_row_df, solution_df], ignore_index=True)
                continue

            hedge_ratio = 1
            ''' Get the target sensitivities for the current rating , then use the invetment value and hedge ratio to generate the final target used in the optimization'''
            target_prep = df_portfolio[df_portfolio['rating'] == rating].drop(df_portfolio.columns[[0, 1]], axis=1)

            target = (target_prep) / investment_val
            target = target.to_numpy()[0]
            target = target.T * -1

            """ For Test Purposes (target krds without dividing by investment values)"""
            krd_target = (target_prep * hedge_ratio)
            krd_target = krd_target.to_numpy()[0]
            krd_target = krd_target.T * -1

            ''' The target sensitivities for provincial assets in the total portfolio are the remainder from the total target minus the solved sensitivities for the other bond ratings '''
            if (portfolio == 'Total') & (rating == 'Provincial'):

                ''' first calculate the solved dollar sensitivities for all the other ratings by multiplying the solved weights by the krds and the investment value '''
                for bond_rating in ['corporateBBB', 'Federal', 'corporateAAA_AA', 'corporateA']:
                    ''' Getting the KRDs '''
                    bond_krd = KRDs[KRDs['rating'] == bond_rating]
                    bond_krd = bond_krd.reset_index().drop(bond_krd.columns[[0, 1]], axis=1).drop('index',
                                                                                                  axis=1).to_numpy()

                    ''' Getting the solved weights from solution_df '''  # TODO! could have issue, solved weights is wrong, step that matters
                    weights = solution_df.loc[
                        (solution_df['portfolio'] == 'Total') & (solution_df['rating'] == bond_rating)].drop(
                        columns={'portfolio', 'rating'}).to_numpy().reshape(-1, 1)

                    ''' Getting the total amount allocated for the rating '''
                    investment = Asset_mix['Total'].loc[bond_rating] / 10000

                    ''' calculating the solved sensitivity then adding it as a column in the solved sensitivity df '''
                    solved_sensitivity = np.dot(bond_krd, weights) * investment
                    solved_dollar_sensitivities[bond_rating] = solved_sensitivity[:, 0]

                ''' The sum of all columns is subtracted from the total target sensitivities to obtain the provincial target sensitivities '''
                solved_dollar_sensitivities['Total - prov'] = solved_dollar_sensitivities.sum(axis=1)
                target_prep = df_portfolio[df_portfolio['rating'] == 'Total'].drop(df_portfolio.columns[[0, 1]], axis=1)
                target_prep = target_prep.iloc[0].reset_index(drop=True) + solved_dollar_sensitivities['Total - prov']

                ''' Applying the same procedure to the target like all previous sensitivity targets - 
                multiplying by the hedge ratio (95%) and dividing by the investment value, then transposing'''
                target = (target_prep * hedge_ratio) / investment_val
                target = target.to_numpy()
                target = target * -1

                """ For Test Purposes (target krds without dividing by investment values)"""
                krd_target = (target_prep * hedge_ratio)
                krd_target = krd_target.to_numpy()
                krd_target = krd_target.T * -1

            # for testing
            krd_targets[f"{portfolio}_{rating}"] = krd_target
            # print(krd_targets)
            # print(asset_type)

            ''' the objective of the optimizer is to minimize the difference 
            between the target sensitivities and the calculated sensitivities'''

            def objective(x):
                c = np.multiply(x, krd)
                temp = c.sum(axis=1) - target
                return np.dot(temp.T, temp)

            ''' for the total portfolio, the objective is to maximize expected return, 
            so it uses a different objective in the optimization'''

            def objective_total(x):
                c = np.multiply(x, expected_return.loc[rating].to_numpy().reshape(1, -1))
                d = -c.sum(axis=1)[0]
                return d

            ''' for corporateBBB bonds - used in the constraints'''
            corpBBBweights = [0.1627, 0.2669, 0.4079, 0.1625]
            corpBBBratios = np.divide(corpBBBweights, corpBBBweights[0])

            ''' Setting constraints for the optimizer - corporateBBB uses 
            different constrants using the ratios calculated above to reduce concentration'''
            if (rating == "corporateBBB") & (asset_type != 'mortgage'):
                cons = ({'type': 'eq', 'fun': lambda x: sum(sum(np.multiply(x, krd))) - sum(target)},
                        {'type': 'eq', 'fun': lambda x: np.sum(x) - 1},
                        {'type': 'eq', 'fun': lambda x: x[3] - corpBBBratios[1] * x[2]},
                        {'type': 'eq', 'fun': lambda x: x[4] - corpBBBratios[2] * x[2]},
                        {'type': 'eq', 'fun': lambda x: x[5] - corpBBBratios[3] * x[2]})
            else:
                cons = ({'type': 'eq', 'fun': lambda x: sum(sum(np.multiply(x, krd))) - sum(target)},
                        {'type': 'eq', 'fun': lambda x: np.sum(x) - 1})

            x0 = [1, 0, 0, 0, 0, 0]

            ''' Setting the boundaries for the optimizer this varies based on portfolio and bond rating '''
            if (asset_type == 'public') & ((portfolio == 'ul') or (portfolio == 'np')):
                ''' Universal and Nonpar are allowed to hold negative amounts for buckets one and 2, the exact amount is calculated using the IFE Estimates file'''
                bnds = calc_bounds(given_date, portfolio, sum(Asset_mix[portfolio]))
            elif (portfolio == 'Total') & (rating != 'corporateAAA_AA'):
                ''' For the Total, the bounds used are based on the solved amounts. The sum of the solved amounts for each portfolio is used as a starting point for the remainder of the total to be optimized'''
                bnds = []

                total_bnds = helpers.get_bnds_for_total(Asset_mix,
                                                        solution_df)
                for x in total_bnds.loc[rating]:
                    bnds.append([x, 1])
            elif (rating == "corporateAAA_AA"):
                ''' No corporateAAA_AA bonds in buckets 3 and 4 and 6, so bounds are set to zero for those buckets '''
                bnds = ((0, 1), (0, 1), (0, 0), (0, 0), (0, 1), (0, 0))
                x0 = [1, 0, 0, 0, 0, 0]
            else:
                bnds = [[0, 1], [0, 1], [0, 1], [0, 1], [0, 1], [0, 1]]

            if portfolio == 'Total':
                ''' Uses a different x0 because [0, 0, 0, 0, 0, 1] is sometimes out of bounds and it gives an incorrect solution '''
                sumofbnds = 1 - bnds[0][0] - bnds[1][0] - bnds[2][0] - bnds[3][0] - bnds[4][0] - bnds[5][0]
                x0 = [bnds[0][0], bnds[1][0], bnds[2][0], bnds[3][0], bnds[4][0], bnds[5][0] + sumofbnds]
                # solution = minimize(objective, x0, method='SLSQP', bounds=bnds, constraints=cons)
                solution = minimize(objective_total, x0, method='SLSQP', bounds=bnds, constraints=cons)
            else:
                solution = minimize(objective, x0, method='SLSQP', bounds=bnds, constraints=cons)

            if solution.success:
                misc.log('Successful optimization for ' + rating + ' bonds in ' + portfolio, LOGFILE)

            ''' Append the solved weights to the solution_df '''
            new_row_df = pd.DataFrame(solution.x).T
            new_row_df['portfolio'] = portfolio
            new_row_df['rating'] = rating
            solution_df = pd.concat([new_row_df, solution_df], ignore_index=True)

    # Test output for krd targets
    if asset_type == 'public':
        krd_targets.to_clipboard()

    ''' Create the liability table using the results of the optimization, and add it to the end of the solution_df'''
    liabilities = helpers.liabilities_table(Asset_mix, solution_df)
    solution_df = pd.concat([solution_df, liabilities], ignore_index=True)

    ''' repeat for the surplus table, append to the end of the solution_df'''
    surplus_table = helpers.surplus_table(Asset_mix, solution_df)
    solution_df = pd.concat([solution_df, surplus_table], ignore_index=True)

    ''' Rounds the solution to 4 decimals'''
    solution_df.iloc[:, :6] = solution_df.iloc[:, :6].astype(float).round(5)  # 4

    return solution_df


"""
Functions that read in input (see comments), supporting optimization functions.
"""
# --- input files ---

def get_expected_returns() -> pd.DataFrame:
    """
    This function is used to read in the expected returns, which are maximized for the total portfolio in the optimization process.

    Expected returns; used in optimization function, where expected returns are optimized for the 'Total' portfolio solution.

    See:
    ''' for the total portfolio, the objective is to maximize expected return,
            so it uses a different objective in the optimization'''

    This function reads and interpolates expected bond returns from the "Parallel_tilt_curve_history.xlsx" file for various bond
    ratings and maturity terms.

    Returns:
    pd.DataFrame: A DataFrame containing the interpolated expected returns for different bond ratings and term assumptions.
    """
    file_name = "Parallel_tilt_curve_history.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    expected_returns = pd.DataFrame()
    ratings = ['Federal', 'Provincial', 'corporateAAA_AA', 'corporateA', 'corporateBBB']

    # Read the expected return data for each bond rating
    for sheet in ['analysis_quarterly_RF', 'analysis_quarterly_prov', 'analysis_quarterly_AA', 'analysis_quarterly_A',
                  'analysis_quarterly_BBB']:
        rownum = 27 if sheet == 'analysis_quarterly_RF' else 22
        ws = workbook[sheet]
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        returns = df.loc[rownum:rownum, 'term1':'term30']
        expected_returns = pd.concat([expected_returns, returns], ignore_index=True)

    # Assign bond ratings to the expected returns DataFrame
    expected_returns['ratings'] = ratings
    expected_returns.set_index('ratings', inplace=True)

    # Term assumptions for interpolation
    term_assumptions = [2, 7, 12, 17, 23, 29]
    return_assumptions = pd.DataFrame(columns=[0, 1, 2, 3, 4, 5])


    x = [1, 2, 3, 4, 5, 7, 10, 20, 30]

    # Interpolate expected returns for each rating across terms
    for rating in ratings:
        y = expected_returns.loc[rating].to_numpy()
        temp = interpolate.splrep(x, y, s=0)
        xnew = np.arange(1, 31)
        ynew = interpolate.splev(xnew, temp, der=0)
        return_assumptions.loc[rating] = ynew[term_assumptions]

    return return_assumptions / 100  # Convert to percentage returns

def calc_bounds(given_date: dt, portfolio: str, total: float) -> List[List[float]]:
    """
    Calculates the optimization bounds for cashflow buckets based on historical data for a given portfolio.

    Parameters:
    given_date (datetime): The date to base the historical data on.
    portfolio (str): The portfolio type (such as 'ul' or 'np') for which bounds are being calculated.
    total (float): The total portfolio value to normalize bounds.

    Returns:
    List[List[float]]: A list of bounds for each cashflow bucket, defining the lower and upper limits.
    """
    if portfolio not in ['ul', 'np']:
        return [[0, 1]] * 6  # Default bounds for other portfolios

    year = given_date.strftime('%Y')
    year_folder = given_date.strftime('%Y')
    quarter = ((given_date.month - 1) // 3) + 1
    prev_quarter = quarter - 1
    if prev_quarter == 0:
        prev_quarter = 4
        year = str(given_date.year - 1)

    quarter = str(quarter)
    prev_quarter = str(prev_quarter)

    # Construct the file name based on the quarter and year
    if given_date.year == 2024 and quarter == '1':
        file_name = f"{portfolio} IFE Estimate Q1 2024.xlsx"
    else:
        file_name = f"{portfolio} IFE Estimate Q{quarter} {year}.xlsx"

    path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation", year_folder,
                              'IFE estimates', f'Q{quarter}', file_name)
    try:
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)
    except FileNotFoundError:
        file_name = f"{portfolio} IFE Estimate Q{prev_quarter} {year} to Q{quarter}.xlsx"
        path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation",
                                  year_folder, 'IFE estimates', f'Q{quarter}', file_name)
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    ws = workbook['CF']
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    # Retrieve the present value (PV) for each cashflow bucket
    cf_pvs = df.iloc[1:7, 34].tolist()
    bounds = []

    # Define bounds for each bucket, allowing short positions if PV is negative
    for pv in cf_pvs:
        if pv >= 0:
            bounds.append([0, 6])
        else:
            bounds.append([pv / total, 6])

    return bounds

