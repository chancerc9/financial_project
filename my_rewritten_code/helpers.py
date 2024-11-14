"""
Name: .py

Purpose:

Functions:

Side effects:

"""


# Standard library imports
import argparse
import datetime as datetime
import json
import os
import sys
from collections import OrderedDict
from typing import Dict, Any

# Third-party library imports
import numpy as np
import openpyxl
import pandas as pd
import psycopg2
from dateutil.relativedelta import relativedelta
from psycopg2.extras import DictCursor
from scipy import interpolate
from scipy.optimize import minimize

# Project-specific imports
from equitable.chronos import conversions, offsets
from equitable.db.db_functions import execute_table_query
from equitable.db.psyw import SmartDB
from equitable.infrastructure import jobs, sendemail, sysenv

# Required custom modules
import file_utils
import model_portfolio_process as mp

# Pandas configuration
pd.set_option('display.width', 150)

# Add system paths
sys.path.append(sysenv.get("ALM_DIR"))  # Add ALM_DIR to system path for additional modules

# Database connections
BM_conn = SmartDB('Benchmark')
BM_cur = BM_conn.con.cursor()

Bond_conn = SmartDB('Bond')
Bond_cur = Bond_conn.con.cursor()

General_conn = SmartDB('General')
General_cur = General_conn.con.cursor()


### CODE ###

#class GetData:


#class CreateShocks:


## Beginning of code ##

# 1. Get bond curves and process (add half-year intervals, choose) - replaces get_bond_curves() and - functionality


def get_bond_curves(GivenDate: datetime) -> pd.DataFrame:
    """
    Retrieves and processes bond curves from the database for a given date. Bond curves remain annual curves as current.

    Parameters:
    GivenDate (datetime): The date for which bond curves are retrieved.

    Returns:
    pd.DataFrame: A DataFrame containing processed bond curves with selected bond types.
    """
    # SQL query to retrieve bond curve data from the database for the given date
    get_bond_curves_query = f"""
                    SELECT *
                    FROM bondcurves_byterm
                    WHERE date= '{GivenDate.date()}' 
                    """
    
    # Query to retrieve column names from the bond curves table
    get_column_names_query = '''SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'bondcurves_byterm';'''
    col_names = [name[0] for name in execute_table_query(get_column_names_query, 'General', fetch=True)]

    # Execute the query and create a DataFrame with the result
    df = pd.DataFrame(execute_table_query(get_bond_curves_query, 'General', fetch=True), columns=col_names)

    # Clean the DataFrame by removing unnecessary columns and transposing it
    df.set_index('name', inplace=True)
    df.drop(['date', 'curvetype'], axis=1, inplace=True)
    df = df.transpose()
    df.reset_index(inplace=True)

    # Select and rename specific bond categories for further analysis
    df = df[['CANADA', 'Provincial', 'AAA & AA', 'A', 'BBB', 'Corporate']]
    df.rename(columns={'CANADA': 'Federal', 'AAA & AA': "CorporateAAA_AA", 'A': 'CorporateA', 'BBB': 'CorporateBBB'}, inplace=True)

    # Shift and divide by 100 to normalize the rates
    df = df.shift()[1:]
    df = df / 100

    return df  # Returns df: a Dataframe of bond curves for all years, per annum (IIRC)


def create_bucketing_table() -> pd.DataFrame:
    """
    Creates a bucketing table with term intervals.

    Returns:
    pd.DataFrame: A DataFrame with term buckets and their respective lower and upper bounds.
    """
    # Create a DataFrame with term buckets ranging from 0.5 to 35 years (70 intervals)
    d = {'Term': list(np.linspace(start=0.5, stop=35, num=70))}
    df = pd.DataFrame(data=d)

    # Calculate the lower and upper bounds for each bucket
    df['Lower_Bound'] = (df['Term'] + df['Term'].shift(1)) / 2  # TODO!  LMAO this is equal to calculating the +25, -25 lower_bound and upper_bound that other functions implement - and may be faster but the others could be considered more intuitive - ask google search or MS Copilot for this terminology or to double check
    df['Upper_Bound'] = df['Lower_Bound'].shift(-1)
    
    # Adjust the first and last bounds
    df.iloc[0, 1] = 0
    df.iloc[-1, 2] = 100 # TODO! should the last bound be 100? EDIT: YES! Can revert to prior git version

    return df


def create_weight_tables(ftse_data: pd.DataFrame):
    """
    Creates weight tables for each bond rating based on subindex percentages.

    Usage:
    Used by create_KRD_tables to aggregate bonds into 6 buckets.

    Parameters:
    ftse_data (pd.DataFrame): A DataFrame containing bond information from the FTSE universe.

    Returns:
    weight_dict (Dict[str, pd.DataFrame]): A dictionary of weight tables for each bond rating.
    total_universe_weights (pd.DataFrame): A DataFrame with total market weights for each rating and term bucket.
    """
    buckets = [1, 5.75, 10.75, 15.75, 20.75, 27.75, 35.25]  # Predefined term buckets
    weight_dict = {}

    total_universe_weights = pd.DataFrame(
        index=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        columns=list(range(1, 7)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        column_to_look_in = "RatingBucket" if rating != 'Corporate' else "Sector"  # Revised to be more concise and clear - Brenda
        
        # Create bucketing table
        df = create_bucketing_table()  # This is for the 70 buckets; weights is for the 6 buckets (from 70 buckets)
        
         # Sum market weights within each bucket
        for x in range(6):
            lower_b = buckets[x]
            upper_b = buckets[x + 1]
            column_name = f"{lower_b} - {upper_b}"

            # Calculate total market weight for the given rating and term bucket
            df[column_name] = df.apply(lambda row: ftse_data.loc[
                (ftse_data[column_to_look_in] == rating) &
                (ftse_data['TermPt'] < upper_b) & # if between lower and upper bounds && between the Lower and Upper bouds by create bucketing table - curious if somehow the cashflows were not affected, because the changes were small and only builds in KRDs?
                (ftse_data['TermPt'] >= lower_b) & # should make upper_b for ftse data as = else isn't as accurate imo
                (ftse_data['TermPt'] < row['Upper_Bound']) &
                (ftse_data['TermPt'] > row['Lower_Bound'] - 0.0001)
            ]['marketweight_noREITs'].sum(), axis=1)

            total_universe_weights.loc[rating, x + 1] = df[column_name].sum()
            
            # Dividing by the sum of the column to get the weight as a percentage of the subindex; ie,
            # Normalize by the sum of market weights
            df[column_name] = df[column_name] / df[column_name].sum()

        weight_dict[rating] = df
        # weight_dict[rating] = df.fillna(0) # TODO: I added this fillNaN

    return weight_dict, total_universe_weights


def create_general_shock_table() -> pd.DataFrame:
    """
    Creates a general shock table to calculate shocks for each security type.

    Returns:
    pd.DataFrame: A DataFrame containing the shock values for different term buckets.
    """
    shock_size = 0.0001
    buckets = [0, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30, 100]
    
    # Create a DataFrame for shocks with 70 term intervals
    df = pd.DataFrame(columns=buckets, index=list(np.linspace(start=0.5, stop=35, num=70)))
    df[0] = df.index  # Initialize the first column with term points
    
    # Apply shock formula for each bucket
    for i in range(1, 11):
        df[buckets[i]] = df.apply(lambda row: (
            (1 - (row[0] - buckets[i]) / (buckets[i + 1] - buckets[i])) * shock_size) 
            if (buckets[i] <= row[0] <= buckets[i + 1]) 
            else (((row[0] - buckets[i - 1]) / (buckets[i] - buckets[i - 1]) * shock_size) 
                  if buckets[i - 1] <= row[0] <= buckets[i] else 0), axis=1)

    # TODO: temp code to overwrite shock tables
    # Assuming df and buckets are already defined

    for i in range(0, 1):
        df[buckets[1]].iloc[i] = shock_size

    for i in range(60, 70): # df[buckets[10]].iloc[69]  is last one ; i = 70 is out of bounds
        df[buckets[10]].iloc[i] = shock_size

    # Loop over the specified values of i
    #for i in [1, 30]:
        # Access the column df[buckets[i]]
    #    column_name = buckets[i]

        # Modify the specified ranges of rows
    #    for j in range(60, 69):
    #        df.loc[j, column_name] = shock_size  # Using .loc to access by label/index

     #   for j in range(0, 1):
     #       df.loc[j, column_name] = shock_size  # Using .loc to access by label/index

    # TODO! end
    # Drop the last bucket (100) as it is not needed
    df = df.drop(100, axis=1)

    # TODO! Debugging_steps - general shock table - date is not even needed, since aply same for all dates
    # can write this one to EXCEL

    
    return df

  #"""
  #Interpolating for half-years - side-eff 1
  #"""

# TODO: temp function, can split up functionality

def create_semi_annual_bond_curves(curves) -> pd.DataFrame: # curves from ftse curves
    # Interpolating bond curves for half-year intervals (linear interpolation; take average of up-down years)
    curves_mod = pd.DataFrame(
        columns=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        index=list(np.linspace(start=0.5, stop=35, num=70)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for i in range(1, 35):
            curves_mod.loc[i, rating] = curves.loc[i, rating]
        for yr in np.linspace(start=1.5, stop=34.5, num=34):
            curves_mod.loc[yr, rating] = (curves.loc[yr + .5, rating] + curves.loc[yr - .5, rating]) / 2

    curves_mod.loc[0.5] = curves_mod.loc[1]
    curves_mod.fillna(method='ffill', inplace=True)

    return curves_mod


# Does NOT modify bond curves parameter; assigns to new value from reference data by dereferencing value

# Applies the shocks to the bond curves for each rating and store results in shocks_dict - side eff 2 (main eff...major purpose)
def create_shock_tables(curves, GivenDate: datetime) -> dict[str, Any]: # CURVES are BOND CURVES (not yet interpolated, just from get_bond_curves) - would help with objectify here so we know what it needs to be can decouple and use specific forced fn attr method etc from bond curves TODO!
    """
    Function to create up and down shock tables for bond curves
    """
    # makes a dictionary containing tables for up shocks and down shocks for each rating
    shocks_dict = {}
    up_shocks = create_general_shock_table() # creates a df with col named '0', '1', etc
    # """old code


    cur_date = GivenDate.strftime('%Y%m%d')

    folder_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), 'Benchmarking', 'Test', 'benchmarking_outputs',
                               'Brenda', 'shocks_table')

    excel_filename = 'shocks_table'
    file_path = os.path.join(folder_path, f'{excel_filename}_{cur_date}.xlsx')

    if not os.path.exists(folder_path):
        os.mkdir(folder_path)

    if not os.path.exists(file_path):
        with pd.ExcelWriter(file_path) as writer:
            up_shocks.to_excel(writer, sheet_name='general_shocks')
    else:
        print('shocks file for this quarter already exists - cant make a file with the same name')

    # can use general writer for this - adopt this into general writer (this works already as one sheet so use this code)

    down_shocks = create_general_shock_table() # TODO: lol
    down_shocks = -down_shocks # can decouple into classes
    down_shocks[0] = -down_shocks[0] # this column (called '0' lmao) holds the bucket numbers, such a weird df... it has both indices and col with dupe nums, guess this fn expects that
    # end of old code """

    #### Interpolates BOND CURVES for half years

    ## (*begin) Interpolates half-year curves (avg of 1, 2y rate for 1.5 yr curve
        # 1+1/2 is average of 1y and 2y rate)

    # Interpolating bond curves for half-year intervals (linear interpolation; take average of up-down years)
    curves_mod = pd.DataFrame(
        columns=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        index=list(np.linspace(start=0.5, stop=35, num=70)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for i in range(1, 35):
            curves_mod.loc[i, rating] = curves.loc[i, rating]
        for yr in np.linspace(start=1.5, stop=34.5, num=34):
            curves_mod.loc[yr, rating] = (curves.loc[yr + .5, rating] + curves.loc[yr - .5, rating]) / 2

    curves_mod.loc[0.5] = curves_mod.loc[1]
    curves_mod.fillna(method='ffill', inplace=True)

    cur_date = GivenDate.strftime('%Y%m%d') # givendate to str - consider doing it here or as fn

    CURR_DEBUGGING_PATH = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), 'Benchmarking', 'Test',
                                       'benchmarking_outputs', 'Brenda', cur_date, 'Debugging_Steps')
    # CURR_FILE_PATH = os.path.join(CURR_DEBUGGING_PATH, 'ftse_bond_curves.xlsx')
    os.makedirs(CURR_DEBUGGING_PATH, exist_ok=True)
    file_utils.write_results_to_excel(curves_mod, CURR_DEBUGGING_PATH, cur_date, 'interpolated_bond_curves')

    #### (*end)

    # Apply up and down shocks to bond curves
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for direction in ['Up', 'Down']:
            table_name = rating + ' - ' + direction
            if direction == 'Up':
                df = up_shocks
            else:
                df = down_shocks

            # Add shocks to bond curves
            df = df.add(curves_mod[rating], axis=0)
            df[0] = curves_mod[rating]

            # Apply power function to simulate the bond curve transformation after shocks
            df += 1
            df['Powers'] = df.index
            df = df.pow(df['Powers'], axis=0)
            df = 1 / df
            df = df.drop('Powers', axis=1)

            shocks_dict[table_name] = df

    return shocks_dict

# (*begin) Takes each year and looks at rating and FTSE universe (half-year would be from .25 to .75; up quarter year and down quarter year for half year, and so on for every year

##### version 1-OLD begins #####

# Function to calculate the average coupon rate for a specific bond rating and year
# It uses the FTSE data to filter bonds based on the given rating and term (maturity year).
# The average coupon is weighted by the notional weight of the bond, excluding REITs.
# The 'price' is MVAI (market weighted price after interest), and we divide it out so removed the market weighting to retain the
# Notional weighting

# TODO! This actually buckets it for the cashflows
#  this function is directly related to the create_cashflows_70 function and does the bucketing for it
def calc_avg_coupon(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the average coupon rate for a specific bond rating and year, weighted by the notional weight of the bond.

    Parameters:
    year (float): The specific year (maturity) to calculate the coupon for. Called 'Bucket", a bucket of 0.5 increments from 0.5 ttm to 35 ttm
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The average coupon rate for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) &  # TODO! essentially, the cashflows 70 uses an appropriate bucketing method in calc_pv and calc_avg_coupon and the weights (6) uses create_bucketing_table to determine the 6th bucket weightings from the FTSE universe, which uses an upper bound of 100 - this mainly affects provincial bond sensitivity weightings, as, the provincial bonds comprise more of the ttms > 35.25. notice that this is bucket 0 in ftse universe, already determined - so this function could simplify it further. Moreover, the 70 cashflows use a different weighting system, and, I presume are less sensitive in 6 weightings than the KRD sensitivities - should have a more formalized (or ocnsistent) system of weighting the exact same way, imo
                       (ftse_data['marketweight_noREITs'] > 0)]   # TODO! NOTE: this uses a DIFFERENT bucketing system than create_bucketing_tables() which is used for the bounds of calculating the 6 weights from 70 tables. Lol, this is funny

    # If no bonds match the criteria, return a coupon rate of 0
    if df.empty:
        return 0

    # Otherwise, calculate the weighted average coupon rate, dividing by 2 for semi-annual coupon payments. As follows:
        # 1. Multiply the market weight by the coupon rate and divide by the market value-adjusted interest (mvai).
        # 2. Divide the sum of these weighted values by the sum of market weights/mvai.
    avg_coupon = ((df['marketweight_noREITs'] * df['annualcouponrate'] / df['mvai']).sum() / # Change in code to use the price (SAME as EXCEL) ****
                  (df['marketweight_noREITs'] / df['mvai']).sum()) / 2  # Divide by 2 to account for semi-annual coupons
    # second SUMPROD is notional weighting
    # Return the calculated average coupon rate for the given rating and year (average was 0 if no matching bonds were found from FTSE bond databank)

    return avg_coupon

##### version 1-OLD ends #####

""" Notional Weighting """

# Function to calculate the average coupon rate for a specific bond rating and year
# It uses the FTSE data to filter bonds based on the given rating and term (maturity year).
# The average coupon is weighted by the notional weight of the bond, excluding REITs.
# Average coupon is normalized by balancing one SUMPRODUCT against another. (Can delete)
def calc_avg_coupon_market_weight(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the average coupon rate for a specific bond rating and year, weighted by the market weight of the bond.

    Parameters:
    year (float): The specific year (maturity) to calculate the coupon for.
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The average coupon rate for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    """ can also use a filter (boolean mask) here: """

    # In code: LB <= filtered_coupons < UB
    #
    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &  # coupon < UB, coupon >= LB
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) & # >= lower bound ; why not just use >=?
                       (ftse_data['marketweight_noREITs'] > 0)]

    # If no bonds match the criteria, return a coupon rate of 0
    if df.empty:
        return 0

    # (filtered df above)

    weighted_sum = (df['mvai'] * df['marketweight_noREITs'] * (1/df['price'])).sum() # is there an equivalent np.sum() operation?

    weighted_sum_with_coupon = (df['mvai'] * df['marketweight_noREITs'] * (1/df['price']) * df['annualcouponrate']).sum()

    average_coupon = (weighted_sum_with_coupon / weighted_sum) / 2  # accounts for semi-annual coupons

    """
    # Otherwise, calculate the weighted average coupon rate, dividing by 2 for semi-annual coupon payments. As follows:
    # 1. Multiply the market weight by the coupon rate and divide by the market value-adjusted interest (mvai).
    # 2. Divide the sum of these weighted values by the sum of market weights/mvai.
    avg_coupon = ((df['marketweight_noREITs'] * df['annualcouponrate'] / df[
        'mvai']).sum() /  # Change in code to use the price (SAME as EXCEL) ****
                  (df['marketweight_noREITs'] / df['mvai']).sum()) / 2  # Divide by 2 to account for semi-annual coupons
    """

    # Return the calculated average coupon rate for the given rating and year (average was 0 if no matching bonds were found from FTSE bond databank)

    return average_coupon

# TODO: Cashflows * interpolated (unshocked) curves

def calc_pv_two(coupons: pd.DataFrame, rating: str, semi_annual_curves: pd.DataFrame): # use generic bond curves for now
    """
    Calculates the present value (PV) of bonds for a specific bond rating and year.

    Parameters:
    year (float): The specific year (maturity) to calculate the PV for.
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The present value (PV) for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    #for i in range(0, 70):

    #coupons.iloc[:, :73]

    # I.e., Calculate the present value (PV) as the weighted sum of market value-adjusted interest (mvai)
    # TODO: change to excel!
    pv = coupons * semi_annual_curves # use array broadcasting, hopefully it applies on all the cols of coupons which by that i mean the cfs (includes principal)

    # Return the calculated present value for the given rating and year
    return pv


# Function to calculate the present value (PV) of bonds for a specific rating and year
# It uses the FTSE data to filter bonds based on the rating and term and then calculates the PV.
def calc_pv(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the present value (PV) of bonds for a specific bond rating and year.

    Parameters:
    year (float): The specific year (maturity) to calculate the PV for.
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The present value (PV) for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) &
                       (ftse_data['marketweight_noREITs'] > 0)]

    # If no bonds match the criteria, return a PV of 0
    if df.empty:
        return 0

    # Otherwise, calculate the present value (PV) by summing up the product of the market weight and the bond's 
    # market value-adjusted interest (mvai), then dividing by the sum of the market weights.
    
    # I.e., Calculate the present value (PV) as the weighted sum of market value-adjusted interest (mvai)
    # TODO: change to excel!
    pv = (df['marketweight_noREITs'] * df['mvai']).sum() / df['marketweight_noREITs'].sum()

    # Return the calculated present value for the given rating and year
    return pv


## USED FUNCTIONs ##

# (*end)

## USED FUNCTION ##
def create_cf_tables(ftse_data): #, GivenDate: pd.Timestamp):
    # uses the average coupon rate to calculate annual cashflows for each rating type
    cf_dict = {}
    years = list(np.linspace(start=0.5, stop=35, num=70))
    buckets = list(np.linspace(start=0.5, stop=35, num=70))
    df = pd.DataFrame(columns=years, index=buckets)
    df.insert(0, 'Bucket', buckets)
    df.insert(1, 'Principal', 100)

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:

        df = pd.DataFrame(columns=years,index=buckets)
        df.insert(0, 'Bucket', buckets)
        df.insert(1, 'Principal', 100)

        df['PV'] = df.apply(lambda row: calc_pv(row['Bucket'], rating, ftse_data), axis=1)
        df['Coupon'] = df.apply(lambda row: calc_avg_coupon(row['Bucket'], rating, ftse_data), axis=1)

        coupons = df.pop(df.columns[-1])
        df.insert(2, 'Coupon', coupons)

        for col in np.linspace(start=0.5, stop=35, num=70):
            df[col] = df.apply(lambda row: row['Coupon'] if row['Bucket'] > col else ((row['Coupon'] + row['Principal']) if row['Bucket'] == col else 0), axis=1)

        cf_dict[rating] = df.iloc[:, :73]
        cf_dict[rating + 'PV'] = df.iloc[:, 73]

    return cf_dict

# Input: ftse_data - a DataFrame containing bond information.
# Output: cf_dict - a dictionary of cashflow tables and their respective present values for each bond rating.

def create_sensitivity_tables(cashflows: Dict[str, pd.DataFrame], shocks: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Calculates cashflow sensitivities based on shocks applied to bond curves.

    Parameters:
    cashflows (Dict[str, pd.DataFrame]): A dictionary containing cashflow tables for each bond rating.
    shocks (Dict[str, pd.DataFrame]): A dictionary containing shock tables for each bond rating.

    Returns:
    Dict[str, pd.DataFrame]: A dictionary of sensitivity tables for each bond rating.
    """
    sensitivities_dict = {}  # Dictionary to store sensitivity tables
    buckets_krd = [0, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30]  # KRD buckets

    # Iterate through each bond rating type to calculate sensitivities
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        # Retrieve cashflows and shocks for the current rating
        cfs = cashflows[rating]  # Retrieve cashflows for the current rating
        ups = shocks[rating + ' - Up']  # Retrieve up shock table for the current rating
        downs = shocks[rating + ' - Down']  # Retrieve down shock table for the current rating

        """
        ## sumproduct for each, changed to get the sensitivities
        # cahsflows for the square (70*70) table, and it fits into the 10*70 sensitivities that it matches up to - sum to each one, cahsflow*shocks.
        """
        # Create empty DataFrames to store sensitivity data for up and down shocks
        df_up = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_up.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        df_down = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_down.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        # Calculate sensitivities by summing the product of cashflows and shocks
        for x in range(1, 11):
            for i in range(70):
                df_up.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * ups.iloc[:, x])  # Multiply cashflows by up shocks
                df_down.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * downs.iloc[:, x])  # Multiply cashflows by down shocks

        # Calculate the average sensitivity (difference between down and up shocks divided by 2)
        average_sensitivity = (df_down - df_up) / 2 * 10000

        # Add bucket information and transpose for better readability
        average_sensitivity['Bucket'] = list(np.linspace(start=.5, stop=35.5, num=71))
        average_sensitivity = average_sensitivity.transpose()
        average_sensitivity = average_sensitivity.drop(70, axis=1)
        average_sensitivity = average_sensitivity.iloc[1:]

        # Insert bucket names for KRD
        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        # TODO: NEW CODE
        avg_sensitivity = average_sensitivity
        # End

        for x in range(10):
            for i in range(70):
                # TODO: for the PV stuff, old code:
                """
                # Safe division, handling division by zero and inf
                numerator = average_sensitivity.iloc[x, i + 1]
                denominator = cashflows[rating + 'PV'].iloc[i]

                # Use np.divide with where clause to avoid division by zero and handle inf
                average_sensitivity.iloc[x, i + 1] = np.divide(numerator, denominator, out=np.zeros_like(numerator), where=denominator != 0)  # Gets the dollar-weighted amounts
                """
                # TODO: NEW CODE
                pv = np.sum(cashflows[rating].iloc[i, 3:] * ups.iloc[:, 0]) # it selects the row, nice (row, which are a bucket) - and ups.iloc[:,0] holds the PV values; of discounted curves
                average_sensitivity.iloc[x, i + 1] = avg_sensitivity.iloc[x, i + 1] / pv
                # End
        # Store the calculated sensitivity table for the rating
        sensitivities_dict[rating] = average_sensitivity

        """ Try this if the above doesn't work:
        # Insert bucket names for KRD and normalize by present values (PV)
        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        # Removed: avg_sensitivity = average_sensitivity - Brenda
        
        # Normalize the sensitivities by the cashflow present values (PV)
        for x in range(10):
            for i in range(70):
                average_sensitivity.iloc[x, i + 1] = average_sensitivity.iloc[x, i + 1] / cashflows[rating + 'PV'].iloc[i]

        # Store the calculated sensitivity table for the rating
        sensitivities_dict[rating] = average_sensitivity
        """

    return sensitivities_dict


"""    # Create empty DataFrames for storing up and down shock sensitivities
        df_up = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_up.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        df_down = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_down.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        # Old:
        #'''
        for x in range(1,11):
            for i in range(70):

                df_up.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * ups.iloc[:, x])  # cashflows table, sum prod
                df_down.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * downs.iloc[:, x])
        #'''
        # New (Brenda):
        #df_up = np.sum(cfs.iloc[:, 3:].values * ups.values[:, 1:], axis=1)
        #df_down = np.sum(cfs.iloc[:, 3:].values * downs.values[:, 1:], axis=1)
        # New (Brenda)

        up_shock_sensitivities = df_up
        down_shock_sensitivities = df_down
        average_sensitivity = (down_shock_sensitivities - up_shock_sensitivities)/2 * 10000

        #average_sensitivity = (average_sensitivity.divide(ups.iloc[:,0]*100) * 10000).iloc[:,1:]
        average_sensitivity['Bucket'] = list(np.linspace(start=.5, stop=35.5, num=71))

        average_sensitivity = average_sensitivity.transpose()

        average_sensitivity = average_sensitivity.drop(70, axis=1)

        average_sensitivity = average_sensitivity.iloc[1:]

        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        avg_sensitivity = average_sensitivity

        for x in range(10):
            for i in range(70):
                average_sensitivity.iloc[x, i + 1] = avg_sensitivity.iloc[x, i + 1] / cashflows[rating + 'PV'].iloc[i]

        sensitivities_dict[rating] = average_sensitivity

    return sensitivities_dict"""
## (*end)

from typing import Dict

import pandas as pd
import numpy as np
from typing import Dict


def make_cashflow_table(weights: Dict[str, pd.DataFrame], cashflows: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Creates a final cashflow allocation table by combining cashflows with market weights for each bond rating.

    Parameters:
    weights (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing market weights for each bond rating and maturity bucket.
    cashflows (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing cashflows for each bond rating and maturity bucket.

    Returns:
    pd.DataFrame: A combined cashflow allocation table for all bond ratings and term buckets.
    """
    Cashflows = {}
    cols = ['rating', 'term', 'bucket1', 'bucket2', 'bucket3', 'bucket4', 'bucket5', 'bucket6']
    terms = cashflows[next(iter(cashflows))].index  # Assuming term intervals as index in cashflows

    # Iterate over each bond rating to calculate cashflow values
    for rating in weights.keys():
        df = pd.DataFrame(columns=cols, index=terms)
        df['term'] = terms  # Assign term intervals to the DataFrame
        df['rating'] = rating  # Set bond rating

        # Extract only the six relevant columns from weights for each term
        weights_six = weights[rating].iloc[:, 2:8].values  # Shape (70, 6)

        # Perform element-wise multiplication across cashflows and weights for each of the 6 final buckets
        cashflow_matrix = cashflows[rating].iloc[:, 1:71].values  # Shape (70, 70)

        # Calculate cashflow allocation by matrix multiplication
        for i in range(6):
            df[f'bucket{i + 1}'] = (cashflow_matrix * weights_six[:, i].reshape(-1, 1)).sum(axis=1)

        Cashflows[rating] = df  # Store cashflow DataFrame in the dictionary

    # Concatenate all rating-specific DataFrames into one final DataFrame
    final_cashflow_df = pd.concat(Cashflows.values(), ignore_index=True)
    final_cashflow_df.fillna(0, inplace=True)  # Replace NaN values with 0

    return final_cashflow_df

    # makes the final KRD table based on sensitivities and market weight and puts it all together in one dataframe
def make_krd_table(weights: Dict[str, pd.DataFrame], sensitivities: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Creates the final Key Rate Duration (KRD) table by combining market weights and cashflow sensitivities for each bond rating.

    Parameters:
    weights (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing market weights for each bond rating and maturity bucket.
    sensitivities (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing sensitivities for each bond rating and maturity bucket.

    Returns:
    pd.DataFrame: A combined KRD table for all bond ratings and term buckets.
    """
    KRDs = {}
    cols = ['rating', 'term', 'bucket1', 'bucket2', 'bucket3', 'bucket4', 'bucket5', 'bucket6']
    buckets = [1, 2, 3, 5, 7, 10, 15, 20, 25, 30]

    # Iterate over each bond rating to calculate KRD values
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        df = pd.DataFrame(columns=cols, index=range(10))
        df['term'] = buckets  # Assign bucket terms to the DataFrame
        df['rating'] = rating  # Set bond rating

        # Calculate KRD by multiplying sensitivities with market weights for each bucket
        for x in range(2, 8):
            df.iloc[:, x] = df.apply(lambda row: (
                sensitivities[rating].loc[sensitivities[rating]['Bucket'] == row['term']].iloc[:, 1:].values[0] *
                weights[rating].iloc[:, (x + 1)]
            ).sum(), axis=1)

        KRDs[rating] = df  # Store KRD DataFrame in the dictionary

    # Concatenate all rating-specific KRD DataFrames into one final DataFrame
    final_krd_df = pd.concat([KRDs['Federal'], KRDs['Provincial'], KRDs['CorporateAAA_AA'], 
                              KRDs['CorporateA'], KRDs['CorporateBBB'], KRDs['Corporate']], ignore_index=True)

    final_krd_df.fillna(0, inplace=True)  # Replace NaN values with 0
    return final_krd_df

# write a fn for middle step

# TODO: Remove!
    # makes the final KRD table based on sensitivities and market weight and puts it all together in one dataframe


def make_CFs_buckets_table(weights: Dict[str, pd.DataFrame],
                           cashflows_seventy: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Creates a cashflow allocation table by combining market weights and cashflows for each bond rating.

    Parameters:
    weights (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing market weights for each bond rating and maturity bucket.
    cashflows_seventy (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing cashflows for each bond rating and maturity bucket.

    Returns:
    pd.DataFrame: A combined cashflow allocation table for all bond ratings and term buckets.
    """
    cashflows_six = {}
    columns = ['rating', 'term', 'bucket1', 'bucket2', 'bucket3', 'bucket4', 'bucket5', 'bucket6']  # Final 6 buckets
    terms = np.arange(0.5, 35.5, 0.5)  # 70-term buckets

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        df = pd.DataFrame(columns=columns, index=range(len(terms)))
        df['term'] = terms  # Assign bucket terms to the DataFrame
        df['rating'] = rating  # Set bond rating

        # Select only the columns for the 6 final buckets from weights
        weights_six = weights[rating].iloc[:, 2:8].values  # Shape should be (70, 6)

        # Perform element-wise multiplication and summing across terms to allocate cashflows
        for i in range(6):
            # Cashflow and weight multiplication for each bucket
            df[f'bucket{i + 1}'] = np.sum(
                cashflows_seventy[rating].iloc[:, 1:71].values * weights_six[:, i].reshape(-1, 1), axis=1)

        cashflows_six[rating] = df  # Store DataFrame for each rating

    # Concatenate all rating-specific DataFrames into one final DataFrame
    final_cashflow_df = pd.concat(
        [cashflows_six['Federal'], cashflows_six['Provincial'], cashflows_six['CorporateAAA_AA'],
         cashflows_six['CorporateA'], cashflows_six['CorporateBBB'], cashflows_six['Corporate']], ignore_index=True)

    final_cashflow_df.fillna(0, inplace=True)  # Replace NaN values with 0
    return final_cashflow_df

# TODO: end of remove!

# def get_bucket_size(weights: Dict[str, pd.DataFrame], cashflows: Dict[str, pd.DataFrame]) -> pd.DataFrame:


import os
from datetime import datetime



import pandas as pd
import numpy as np
from scipy import interpolate
import openpyxl

def get_expected_returns() -> pd.DataFrame:
    """
    Reads and interpolates expected bond returns from the "Parallel_tilt_curve_history.xlsx" file for various bond ratings and maturity terms.

    Returns:
    pd.DataFrame: A DataFrame containing the interpolated expected returns for different bond ratings and term assumptions.
    """
    file_name = "Parallel_tilt_curve_history.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    expected_returns = pd.DataFrame()
    ratings = ['Federal', 'Provincial', 'corporateAAA_AA', 'corporateA', 'corporateBBB']

    # Read the expected return data for each bond rating
    for sheet in ['analysis_quarterly_RF', 'analysis_quarterly_prov', 'analysis_quarterly_AA', 'analysis_quarterly_A', 'analysis_quarterly_BBB']:
        rownum = 27 if sheet == 'analysis_quarterly_RF' else 22
        ws = workbook[sheet]
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        returns = df.loc[rownum:rownum, 'term1':'term30']
        expected_returns = pd.concat([expected_returns, returns], ignore_index=True)

    # Assign bond ratings to the expected returns DataFrame
    expected_returns['ratings'] = ratings
    expected_returns.set_index('ratings', inplace=True)

    # Term assumptions for interpolation
    term_assumptions = [2, 7, 12, 17, 23, 29]
    return_assumptions = pd.DataFrame(columns=[0, 1, 2, 3, 4, 5])

    x = [1, 2, 3, 4, 5, 7, 10, 20, 30]
    # Interpolate expected returns for each rating across terms
    for rating in ratings:
        y = expected_returns.loc[rating].to_numpy()
        temp = interpolate.splrep(x, y, s=0)
        xnew = np.arange(1, 31)
        ynew = interpolate.splev(xnew, temp, der=0)
        return_assumptions.loc[rating] = ynew[term_assumptions]

    return return_assumptions / 100  # Convert to percentage returns


## New fUNCTION to read in stuff (no need from balance sheet) - check it easier from the balance sheet, easier
# helper function to read in from balance sheet (execel) not from the comapny blcsht its

# so have a function for reading in data
# for modifiability
# yu don't know how much i know the code - which is a lot
# i was doing it durin

# before it was reading in assets from balance sheet, but now its reading in data from excel sheet

# ReadInData
# this class also process it

# i forgot, bond class or calc omg
# recall and reconnect
#check the idea again
# thibk again for what makes sense///what was it again
# what did i think and mention


from datetime import datetime

def BSTotals(given_date: datetime, sheet_version: int) -> dict:
    """
    Retrieves the balance sheet totals from the "SBS Totals.xlsx" file based on the provided date.

    Parameters:
    given_date (datetime): The date for which the balance sheet totals are requested.
    sheet_version (int): Determines if totals or segments are returned (1 for segments, 0 for totals).

    Returns:
    dict: A dictionary containing balance sheet totals for different categories.
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1

    year_quarter = f"{year}Q{quarter}"

    quarter = f"Q{quarter}"


    # file_name = "SBS Totals - Brenda.xlsx"
    file_name = "SBS Totals.xlsx"
    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True)

    # Retrieve the appropriate worksheet based on sheet version
    # ws = workbook[year_quarter] if sheet_version == 1 else workbook[year_quarter + ' (Total)']
    ws = workbook['Segments'] if sheet_version == 1 else workbook['Total']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # Extract and return totals for relevant categories
    totals = {
        'ACCUM': df.loc[2, 'ACCUM'],
        'PAYOUT': df.loc[2, 'PAYOUT'],
        'UNIVERSAL': df.loc[2, 'UNIVERSAL'],
        'NONPAR': df.loc[2, 'NONPAR'],
        'GROUP': df.loc[2, 'GROUP'],
        'PARCSM': df.loc[2, 'PARCSM'],
        'SEGFUNDS': df.loc[2, 'SEGFUNDS'],
        'Surplus': df.loc[2, 'Surplus'],
        'Total': df.loc[2, 'Total']
    } # rewrote BSTotals to reduce unecessary code / overwritten initializaitons

    return totals


def percents(given_date: datetime, curMonthBS: bool = False) -> pd.DataFrame:
    """
    Retrieves asset mix percentages from the "Mix_hardcoded.xlsx" file for the given date.

    Parameters:
    given_date (datetime): The date for which the asset mix percentages are requested.
    curMonthBS (bool): If True, adjusts the quarter to the next one if applicable. Default is False.

    Returns:
    pd.DataFrame: A DataFrame containing asset mix percentages for various bond ratings.
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    if curMonthBS and quarter < 4:
        quarter += 1
    year_quarter = year + "Q" + str(quarter)

    quarter = f"Q{quarter}"

    file_name = "Asset Mix.xlsx"
    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    workbook = openpyxl.load_workbook(path_input, data_only = True) # TODO! new change - linked data vals work
    # ws = workbook[year_quarter]  # i.e., ws = workbook['2024Q1']
    ws = workbook['Sheet1']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns).set_index('rating')
    
    # Initialize surplus and SEGFUNDS columns
    df['Surplus'] = df['SEGFUNDS'] = 0

    # Filter rows to include only relevant bond categories
    df = df.loc[['Federal',
                 'Provincial',
                 'CorpAAA_AA',
                 'CorpA',
                 'CorpBBB',
                 'MortgagesInsured',
                 'MortgagesConv',
                 'PrivateAA',
                 'PrivateA',
                 'PrivateBBB',
                 'PrivateBB_B']]
    
    return df

from typing import Dict

def solution_dollar_amounts(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculates the dollar allocation for each bond rating across different portfolios based on the asset mix and solution data.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations by rating.

    Returns:
    pd.DataFrame: A DataFrame with dollar allocations for each portfolio and rating.
    """
    weights = Asset_mix[['Accum', 'group', 'ul', 'Payout', 'np']].stack().sort_index()  # (*Brenda)
    weights2 = weights.reset_index(drop=True)
    
    # Filter the solution DataFrame to exclude 'Liability' and 'Total' portfolios
    sols = solution_df[(solution_df['portfolio'] != 'Liability') & (solution_df['portfolio'] != 'Total')].set_index(['rating', 'portfolio']).sort_index()  # Python has great order-of-operations (*Brenda)
    sols2 = sols.reset_index(drop=True)
    
    # Calculate weighted dollar allocations
    w = sols2.mul(weights2, axis=0)
    w['rating'] = sols.reset_index()['rating']
    w['portfolio'] = sols.reset_index()['portfolio']
    w = w.set_index(['portfolio', 'rating'])

    # Group by rating and sum for total allocation
    w_grouped = w.groupby('rating')
    for index, row in w_grouped:
        total_values = row.sum()
        total_values['rating'] = index
        total_values['portfolio'] = 'Total'
        w = pd.concat([w, pd.DataFrame(total_values).T.set_index(['portfolio', 'rating'])])

    return w.reset_index()

''' This function takes in the asset mix and the solved solution up to this point to calculate how much of the total allocation has been allocated in each portfolio. Those weights are used as bounds for the total optimization''' # Old comment vs my interpretation (kept both in case err - for now)
def get_bnds_for_total(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculates the bounds for total optimization based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the upper and lower bounds for optimization for each rating.
    """
    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')
    dollars = Asset_mix['Total']
    
    # Calculate bounds by dividing total allocation by asset mix
    bounds = total.div(dollars, axis=0)
    bounds = bounds.where(bounds > 0, 0)
    
    return bounds


def liabilities_table(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a table of liability allocations for each bond rating based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the liability allocations by bond rating.
    """
    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')
    dollars = total.sum(axis=1)
    
    # Calculate liability allocations by dividing by total dollars
    liabilities = total.div(dollars, axis=0)
    liabilities['rating'] = liabilities.index
    liabilities['portfolio'] = 'Liability'
    
    return liabilities.reset_index(drop=True)


def surplus_table(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a table of surplus allocations for each bond rating based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the surplus allocations by bond rating.
    """
    npt_weights = Asset_mix['Total']
    npt_sol = solution_df[solution_df['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')

    # Calculate optimized solution using weights
    optimization_sol = npt_sol.mul(npt_weights, axis=0)

    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')

    # Calculate the surplus by subtracting total from optimization
    total = optimization_sol - total
    dollars = total.sum(axis=1)
    
    # Calculate surplus allocations by dividing by total dollars
    surplus = total.div(dollars, axis=0)
    surplus['rating'] = surplus.index
    surplus['portfolio'] = 'Surplus'
    
    return surplus.reset_index(drop=True)


from typing import List
from datetime import datetime

def calc_bounds(given_date: datetime, portfolio: str, total: float) -> List[List[float]]:
    """
    Calculates the optimization bounds for cashflow buckets based on historical data for a given portfolio.

    Parameters:
    given_date (datetime): The date to base the historical data on.
    portfolio (str): The portfolio type (such as 'ul' or 'np') for which bounds are being calculated.
    total (float): The total portfolio value to normalize bounds.

    Returns:
    List[List[float]]: A list of bounds for each cashflow bucket, defining the lower and upper limits.
    """
    if portfolio not in ['ul', 'np']:
        return [[0, 1]] * 6  # Default bounds for other portfolios

    year = given_date.strftime('%Y')
    year_folder = given_date.strftime('%Y')
    quarter = ((given_date.month - 1) // 3) + 1
    prev_quarter = quarter - 1
    if prev_quarter == 0:
        prev_quarter = 4
        year = str(given_date.year - 1)

    quarter = str(quarter)
    prev_quarter = str(prev_quarter)

    # Construct the file name based on the quarter and year
    if given_date.year == 2024 and quarter == '1':
        file_name = f"{portfolio} IFE Estimate Q1 2024.xlsx"
    else:
        file_name = f"{portfolio} IFE Estimate Q{quarter} {year}.xlsx"

    path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation", year_folder, 
                              'IFE estimates', f'Q{quarter}', file_name)
    try:
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)
    except FileNotFoundError:
        file_name = f"{portfolio} IFE Estimate Q{prev_quarter} {year} to Q{quarter}.xlsx"
        path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation", 
                                  year_folder, 'IFE estimates', f'Q{quarter}', file_name)
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    ws = workbook['CF']
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    # Retrieve the present value (PV) for each cashflow bucket
    cf_pvs = df.iloc[1:7, 34].tolist()
    bounds = []

    # Define bounds for each bucket, allowing short positions if PV is negative
    for pv in cf_pvs:
        if pv >= 0:
            bounds.append([0, 6])
        else:
            bounds.append([pv / total, 6])

    return bounds

''' given a df with a multi-index, portfolio and rating, this function will sum all rows with the same rating, and append the sum to a new row with portfolio 'Total' '''
def get_totals_for_rating(df: pd.DataFrame, reset_index: bool = False) -> pd.DataFrame:
    """
    Summarizes the total values for each rating in the given DataFrame by aggregating portfolios.

    Parameters:
    df (pd.DataFrame): A DataFrame with multi-index of 'portfolio' and 'rating'.
    reset_index (bool): Whether to reset the index after aggregating. Default is False.

    Returns:
    pd.DataFrame: A DataFrame with total values for each rating, with an additional row for portfolio 'Total'.
    """
    print(df)

    df_copy = df.copy()
    df_grouped = df_copy.groupby('rating')

    # Sum all rows with the same rating and append the total row
    for index, row in df_grouped:
        total_values = row.sum()
        total_values['rating'] = index
        total_values['portfolio'] = 'Total'
        total_values_df = pd.DataFrame(total_values).T.set_index(['portfolio', 'rating'])
        df_copy = pd.concat([df_copy, total_values_df])

    return df_copy.reset_index() if reset_index else df_copy


def public_sensitivities(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves public asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for public asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"

    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"
    quarter = f"Q{quarter}"

    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    sheet = 'public'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df


def private_sensitivities(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves private asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for private asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)

    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"

    quarter = f"Q{quarter}"


    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)

    sheet = 'private'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df


def mortgage_sensitivities(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves mortgage asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for mortgage asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    # path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)

    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"

    quarter = f"Q{quarter}"


    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, quarter)
    os.makedirs(dir_path, exist_ok=True)

    path_input = os.path.join(dir_path,
                              file_name)


    sheet = 'mortgage'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df
