"""
Name: datahandler.py

Purpose:
    FTSE queries db; other functions read-in data from excel files.
    MODIFIES DATA

    Allows for data protection via classes
    Provides data protection

Functions:

Side effects:

"""
# Standard library imports
import os
from datetime import datetime
from pathlib import Path
from typing import Tuple

# Third-party imports
import pandas as pd
import numpy as np
import openpyxl


# Local application-specific imports
from equitable.db.db_functions import execute_table_query
from equitable.infrastructure import sysenv


# Read in FTSE data from database
class FTSEDataHandler:
    """
    A class to retrieve, process, and provide controlled access to bond data
    from the FTSE universe.

    The `FTSEDataHandler` class allows users to retrieve FTSE bond data for a
    specified date, process it, and access it in a read-only format through
    a `@property`. Data mutations are internally managed, ensuring that the
    data remains immutable to external code. Users can refresh the data or
    update the date through specific methods.

    Side effects:
        Modifies __data
        Performs query on SQL database - does not modify database.

    Attributes:
    ----------
    given_date : datetime
        The date for which bond data is retrieved.
    data : pd.DataFrame
        A read-only DataFrame containing processed FTSE bond data with additional
        calculated columns.

    Methods:
    -------
    refresh_data():
        Reloads bond data for the current date (i.e., refreshes data without need to change the current date).

    update_date(new_date):
        Updates the date for retrieving bond data and refreshes the data.
    """

    def __init__(self, given_date: datetime):
        """
        Initializes the FTSEDataHandler with a specified date.

        Parameters:
        ----------
        given_date : datetime
            The date for which FTSE bond data will be retrieved and processed.
        Notes:
        ------
        The initial data load occurs upon instantiation by calling `_load_data`.
        """
        self._given_date = given_date
        self._data = None
        self._load_data()

    @property
    def data(self) -> pd.DataFrame:
        """
        Accesses the processed FTSE bond data as a read-only DataFrame.

        Returns:
        -------
        pd.DataFrame
            A copy of the internally stored FTSE bond data DataFrame,
            preventing external modification.

        Notes:
        ------
        The returned DataFrame is a copy of the internal `_data` attribute
        to ensure immutability. Any changes to the returned DataFrame do not
        affect the original data stored within the class.
        """
        return self._data.copy()

    def _load_data(self):
        """
        Retrieves and processes FTSE bond data for the specified date.
        This method executes an SQL query to fetch bond data, processes it to
        calculate additional columns, and stores the processed data in `_data`.

        Notes:
        ------
        This method is called internally by `__init__` and `refresh_data` to
        handle data loading. It applies transformations to add calculated
        columns such as 'marketweight_noREITs', 'Sector', 'RatingBucket',
        'mvai', 'TermPt', and 'bucket' for analysis.
        """
        get_bond_info_query = f"""
            SELECT date, cusip, term, issuername,
            annualcouponrate, maturitydate, yield,
            accruedinterest, modduration, rating,
            industrysector, industrygroup, industrysubgroup,
            marketweight, price
            FROM ftse_universe_constituents
            WHERE date= '{self._given_date.date()}'
        """

        # Execute SQL query and create DataFrame:
        df = pd.DataFrame(execute_table_query(get_bond_info_query, 'Bond', fetch=True))
        df.columns = ['date', 'cusip', 'term', 'issuername', 'annualcouponrate', 'maturitydate', 'yield',
                      'accruedinterest', 'modduration', 'rating', 'industrysector', 'industrygroup',
                      'industrysubgroup', 'marketweight', 'price']

        # --- Process data columns           ---
        #     Effects: modifies _data values ---

        # a) Calculate the market weight excluding real estate (REITs):
        total_marketweight = df['marketweight'].sum()
        real_estate_weight = df[df['industrygroup'] == "Real Estate"]['marketweight'].sum()
        # Variable name == ['market_weight_noREITs']
        df['marketweight_noREITs'] = df.apply(
            lambda row: 0 if row['industrygroup'] == "Real Estate"
            else row['marketweight'] / (total_marketweight - real_estate_weight) * 100,
            axis=1)

        # b) Add classification columns for sector (e.g. a bond name) and rating
        df['Sector'] = df.apply(
            lambda row: row['industrygroup'] if row['industrysector'] == 'Government' else row['industrysector'],
            axis=1)
        # Drop municipal bonds
        df.drop(df[df['Sector'] == 'Municipal'].index, inplace=True)
        df['SectorM'] = df['Sector']
        df['Rating_c'] = df.apply(lambda row: "AAA_AA" if row['rating'] in ['AA', 'AAA'] else row['rating'], axis=1)
        df['RatingBucket'] = df.apply(
            lambda row: row['SectorM'] + row['Rating_c'] if row['SectorM'] == 'Corporate' else row['SectorM'], axis=1)

        # Note: MVAI is accrued interest + price (TODO! can place into a calculations sheet if desired)
        df['mvai'] = df['accruedinterest'] + df['price']

        # c) Calculate term points based on maturity date (TODO! Can isolate float64 365.25 into global var SET_YEAR_LEN)
        df['TermPt'] = df.apply(lambda row: round((row['maturitydate'] - self._given_date.date()).days / 365.25, 2),
                                axis=1)  # TODO! As this is hardcoded, worthwhile putting the calculations on another page - for how we process the ftse data

        # d) Bucket the bonds into six term buckets (conditions => maintainability - *Brenda*) # TODO! remove name
        conditions = [
            (df['TermPt'] < 5.75),
            (df['TermPt'] < 10.75),
            (df['TermPt'] < 15.75),
            (df['TermPt'] < 20.75),
            (df['TermPt'] < 27.75),
            (df['TermPt'] < 35.25)
        ]
        choices = [1, 2, 3, 4, 5,
                   6]  # TODO! Remove comment: # only non-mutating functions methods are for bond terms and curves, not even for ftse data as it manipulates it (!!)
        # For TermPt >= 35.25, bucket = 0
        df['bucket'] = np.select(conditions, choices,
                                 default=0)  # TODO! Remove comment: # np.select() for vectorization (*Brenda* - these comments are removable)
        self._data = df

    def refresh_data(self):
        """
        Reloads FTSE bond data for the current date.
        Notes:
        ------
        This method re-fetches and processes the data based on the current
        `_given_date`, allowing users to refresh the data without modifying
        the specified date.
        """
        self._load_data()

    def update_date(self, new_date: datetime):
        """
        Updates the date and refreshes FTSE bond data.
        Parameters:
        ----------
        new_date : datetime
            The new date for which FTSE bond data should be retrieved.
        Notes:
        ------
        This method sets a new date and automatically triggers data
        re-processing by calling `refresh_data`.
        """
        self._given_date = new_date
        self.refresh_data()



# ----- read in data -----


def get_year_and_quarter(given_date: datetime) -> Tuple[str, int]:
    """
    Extracts the year and quarter from a given date.
    Args:
        given_date (datetime): The date to process.
    Returns:
        Tuple[str, int]: A tuple containing the year as a string
                         and the quarter as an integer (1 to 4).
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    return year, quarter


def set_input_path(given_date: datetime, file_name: str) -> Path:
    """
    Constructs a file path based on the given date, creating the
    necessary directory structure if it does not exist.
    Args:
        given_date (datetime): The date used to determine the folder structure.
        file_name (str): The name of the file to include in the path.
    Returns:
        Path: The full path to the file, including the directory structure.
    """

    # Determine year and quarter
    year, quarter = get_year_and_quarter(given_date)

    # Make directory path
    dir_path = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", "Inputs", year, f"Q{quarter}")
    os.makedirs(dir_path, exist_ok=True)

    # Make file path
    path_input = os.path.join(dir_path,
                              file_name)

    # Return the full file path
    return path_input




### Reads in input ###
"""Reads in assets"""

# Reads in asset totals from SBS_totals.xlsx
def BSTotals(given_date: datetime, sheet_version: int) -> dict:
    """
    Retrieves the balance sheet totals from the "SBS Totals.xlsx" file based on the provided date.

    Parameters:
    given_date (datetime): The date for which the balance sheet totals are requested.
    sheet_version (int): Determines if totals or segments are returned (1 for segments, 0 for totals).

    Returns:
    dict: A dictionary containing balance sheet totals for different categories.
    """

    # file_name = "SBS Totals Modified.xlsx"
    # file_name = "SBS Totals SC.xlsx"
    file_name = "SBS Totals.xlsx"

    # path for given input:
    path_input = set_input_path(given_date, file_name)

    # Open workbook:
    workbook = openpyxl.load_workbook(path_input, data_only=True)

    # Retrieve the appropriate worksheet based on sheet version
    ws = workbook['Segments'] if sheet_version == 1 else workbook['Total']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # Extract and return totals for relevant categories
    totals = {
        'ACCUM': df.loc[2, 'ACCUM'],
        'PAYOUT': df.loc[2, 'PAYOUT'],
        'UNIVERSAL': df.loc[2, 'UNIVERSAL'],
        'NONPAR': df.loc[2, 'NONPAR'],
        'GROUP': df.loc[2, 'GROUP'],
        'PARCSM': df.loc[2, 'PARCSM'],
        'SEGFUNDS': df.loc[2, 'SEGFUNDS'],
        'Surplus': df.loc[2, 'Surplus'],
        'Total': df.loc[2, 'Total']
    }

    return totals


# Asset mix.xlsx reading in function
def percents(given_date: datetime) -> pd.DataFrame:
    """
    Retrieves asset mix percentages from the "Asset Mix.xlsx" file for the given date.

    Parameters:
    given_date (datetime): The date for which the asset mix percentages are requested.

    Returns:
    pd.DataFrame: A DataFrame containing asset mix percentages for various bond ratings.
    """

    file_name = "Asset Mix.xlsx"
    # file_name = "Asset Mix Modified.xlsx"
    # file_name = "Asset Mix SC.xlsx"
    # file_name = "Asset Mix Prov -1%.xlsx"

    path_input = set_input_path(given_date, file_name)

    # Open workbook:
    workbook = openpyxl.load_workbook(path_input, data_only=True)

    # Retrieve the appropriate worksheet:
    ws = workbook['Sheet1']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns).set_index('rating')

    # Initialize surplus and SEGFUNDS columns
    df['Surplus'] = df['SEGFUNDS'] = 0

    # Filter rows to include only relevant bond categories
    df = df.loc[['Federal',
                 'Provincial',
                 'CorpAAA_AA',
                 'CorpA',
                 'CorpBBB',
                 'MortgagesInsured',
                 'MortgagesConv',
                 'PrivateAA',
                 'PrivateA',
                 'PrivateBBB',
                 'PrivateBB_B']]

    return df
# end of Asset Mix.xlsx reading in function.


"""Read in liability sensitivities"""


# Reads in liabilities from 'Targets by Asset Class.xlsx'
def get_liability_sensitivities(given_date: datetime, liability_type: str = 'public'): # file_path: str,
    """
    example:
        file_path: str = "Targets By Asset Class.xlsx"

    liability_type can be one of
    'public'
    'mortgage'
    'private'
        and matches with asset class.

    returns DataFrame of liability sensitivities for selected asset class.
    """

    file_name = "Targets By Asset Class.xlsx"
    # file_name = "Targets By Asset Class SC.xlsx"

    # file_name = "Targets By Asset Class 100%.xlsx"
    # file_name = "Targets By Asset Class Modified.xlsx"

    path_input = set_input_path(given_date, file_name)

    sheet = liability_type
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df
# end of Targets by Asset Class reading in functions.




"""Read in semi-annual bond curves from excel"""
# Reads in bond curves
def get_bond_curves(GivenDate: datetime) -> pd.DataFrame:
    """
    Returns a dataframe of semi-annual bond curves.
    """
    # Set file name and path, and sheet name
    file_name = "Curves.xlsx"
    path_input = set_input_path(GivenDate, file_name)

    sheet = 'Curves'
    # Load workbook, open sheet, read data values
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values

    # Read columns
    columns = next(data)[0:]

    # Set columns as columns of first row
    df = pd.DataFrame(data, columns=columns)

    # Set index as buckets
    df.set_index(df.columns[0], inplace=True)
    df.index.name = 'Term Bucket'

    return df
