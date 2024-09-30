# Standard library imports
import argparse
import datetime as dt
import json
import os
import sys
from collections import OrderedDict

# Third-party library imports
import numpy as np
import openpyxl
import pandas as pd
import psycopg2
from dateutil.relativedelta import relativedelta
from psycopg2.extras import DictCursor
from scipy import interpolate
from scipy.optimize import minimize

# Project-specific imports
from equitable.chronos import conversions, offsets
from equitable.db.db_functions import execute_table_query
from equitable.db.psyw import SmartDB
from equitable.infrastructure import jobs, sendemail, sysenv

# Pandas configuration
pd.set_option('display.width', 150)

# Add system paths
sys.path.append(sysenv.get("ALM_DIR"))  # Add ALM_DIR to system path for additional modules

# Database connections
BM_conn = SmartDB('Benchmark')
BM_cur = BM_conn.con.cursor()

Bond_conn = SmartDB('Bond')
Bond_cur = Bond_conn.con.cursor()

General_conn = SmartDB('General')
General_cur = General_conn.con.cursor()


# for the (?), double-check through terminology

### CODE ###

#class GetData:


#class CreateShocks:


# 1. Get bond curves and process (add half-year intervals, choose) - replaces get_bond_curves() and - functionality

## Beginning of code ##

def get_bond_curves(GivenDate: datetime) -> pd.DataFrame:
    """
    Retrieves and processes bond curves from the database for a given date. Bond curves remain annual curves as current.

    Parameters:
    GivenDate (datetime): The date for which bond curves are retrieved.

    Returns:
    pd.DataFrame: A DataFrame containing processed bond curves with selected bond types.
    """
    # SQL query to retrieve bond curve data from the database for the given date
    get_bond_curves_query = f"""
                    SELECT *
                    FROM bondcurves_byterm
                    WHERE date= '{GivenDate.date()}' 
                    """
    
    # Query to retrieve column names from the bond curves table
    get_column_names_query = '''SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'bondcurves_byterm';'''
    col_names = [name[0] for name in execute_table_query(get_column_names_query, 'General', fetch=True)]

    # Execute the query and create a DataFrame with the result
    df = pd.DataFrame(execute_table_query(get_bond_curves_query, 'General', fetch=True), columns=col_names)

    # Clean the DataFrame by removing unnecessary columns and transposing it
    df.set_index('name', inplace=True)
    df.drop(['date', 'curvetype'], axis=1, inplace=True)
    df = df.transpose()
    df.reset_index(inplace=True)

    # Select and rename specific bond categories for further analysis
    df = df[['CANADA', 'Provincial', 'AAA & AA', 'A', 'BBB', 'Corporate']]
    df.rename(columns={'CANADA': 'Federal', 'AAA & AA': "CorporateAAA_AA", 'A': 'CorporateA', 'BBB': 'CorporateBBB'}, inplace=True)

    # Shift and divide by 100 to normalize the rates
    df = df.shift()[1:]
    df = df / 100

    return df  # Returns df: a Dataframe of bond curves for all years, per annum (IIRC)


def get_ftse_data(givenDate: datetime) -> pd.DataFrame:
    """
    Retrieves bond data from the FTSE universe for a given date and processes it.

    Parameters:
    givenDate (datetime): The date for which bond data is retrieved.

    Returns:
    pd.DataFrame: A DataFrame containing processed FTSE bond data with additional calculated columns.
    """
    # SQL query to retrieve bond information from the FTSE universe
    get_bond_info_query = f"""
                    SELECT date, cusip, term, issuername, 
                    annualcouponrate, maturitydate, yield, 
                    accruedinterest, modduration, rating, 
                    industrysector, industrygroup, industrysubgroup, 
                    marketweight, price
                    FROM ftse_universe_constituents
                    WHERE date= '{givenDate.date()}'
                    """
    
    # Execute the query and create a DataFrame with the result
    df = pd.DataFrame(execute_table_query(get_bond_info_query, 'Bond', fetch=True))
    df.columns = ['date', 'cusip', 'term', 'issuername', 'annualcouponrate', 'maturitydate', 'yield', 
                  'accruedinterest', 'modduration', 'rating', 'industrysector', 'industrygroup', 
                  'industrysubgroup', 'marketweight', 'price']

    # Calculate the market weight excluding real estate (REITs)
    total_marketweight = df['marketweight'].sum()
    real_estate_weight = df[df['industrygroup'] == "Real Estate"]['marketweight'].sum()
    df['marketweight_noREITs'] = df.apply(lambda row: 0 if row['industrygroup'] == "Real Estate" 
                                          else row['marketweight'] / (total_marketweight - real_estate_weight) * 100, axis=1)

    # Add classification columns for sector (e.g. a bond name) and rating
    df['Sector'] = df.apply(lambda row: row['industrygroup'] if row['industrysector'] == 'Government' else row['industrysector'], axis=1)
    df.drop(df[df['Sector'] == 'Municipal'].index, inplace=True)  # Drop municipal bonds
    df['SectorM'] = df['Sector']
    df['Rating_c'] = df.apply(lambda row: "AAA_AA" if row['rating'] in ['AA', 'AAA'] else row['rating'], axis=1)
    df['RatingBucket'] = df.apply(lambda row: row['SectorM'] + row['Rating_c'] if row['SectorM'] == 'Corporate' else row['SectorM'], axis=1)
    df['mvai'] = df['accruedinterest'] + df['price']

    # Calculate term points based on maturity date
    df['TermPt'] = df.apply(lambda row: round((row['maturitydate'] - givenDate.date()).days / 365.25, 2), axis=1)

    # Bucket the bonds into six term buckets (conditions => maintainability - *Brenda*)
    conditions = [
    (df['TermPt'] < 5.75),
    (df['TermPt'] < 10.75),
    (df['TermPt'] < 15.75),
    (df['TermPt'] < 20.75),
    (df['TermPt'] < 27.75),
    (df['TermPt'] < 35.25)
    ]
    choices = [1, 2, 3, 4, 5, 6]
    df['bucket'] = np.select(conditions, choices, default=0)  # np.select() for vectorization (*Brenda* - these comments are removable)

    """
    df['bucket'] = df.apply(lambda row: 1 if row['TermPt'] < 5.75 
                            else (2 if row['TermPt'] < 10.75 
                                  else (3 if row['TermPt'] < 15.75 
                                        else (4 if row['TermPt'] < 20.75 
                                              else (5 if row['TermPt'] < 27.75 
                                                    else 6)))), axis=1)
    """
    return df
  
    # code takes it, puts it into 70 buckets, figure out the coupons and the weights, and puts it back down to 6 buckts (*a) , to find assets to invest, and to match up to our sensitivities
def create_bucketing_table() -> pd.DataFrame:
    """
    Creates a bucketing table with term intervals.

    Returns:
    pd.DataFrame: A DataFrame with term buckets and their respective lower and upper bounds.
    """
    # Create a DataFrame with term buckets ranging from 0.5 to 35 years (70 intervals)
    d = {'Term': list(np.linspace(start=0.5, stop=35, num=70))}
    df = pd.DataFrame(data=d)

    # Calculate the lower and upper bounds for each bucket
    df['Lower_Bound'] = (df['Term'] + df['Term'].shift(1)) / 2
    df['Upper_Bound'] = df['Lower_Bound'].shift(-1)
    
    # Adjust the first and last bounds
    df.iloc[0, 1] = 0
    df.iloc[-1, 2] = 100

    return df


def create_weight_tables(ftse_data: pd.DataFrame):
    """
    Creates weight tables for each bond rating based on subindex percentages.

    Parameters:
    ftse_data (pd.DataFrame): A DataFrame containing bond information from the FTSE universe.

    Returns:
    weight_dict (Dict[str, pd.DataFrame]): A dictionary of weight tables for each bond rating.
    total_universe_weights (pd.DataFrame): A DataFrame with total market weights for each rating and term bucket.
    """
    buckets = [1, 5.75, 10.75, 15.75, 20.75, 27.75, 35.25]  # Predefined term buckets
    weight_dict = {}

    total_universe_weights = pd.DataFrame(
        index=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        columns=list(range(1, 7)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        column_to_look_in = "RatingBucket" if rating != 'Corporate' else "Sector"  # (*Brenda: more concise and clear - simple)
        
        # Create bucketing table
        df = create_bucketing_table()
        
         # Sum market weights within each bucket
        for x in range(6):
            lower_b = buckets[x]
            upper_b = buckets[x + 1]
            column_name = f"{lower_b} - {upper_b}"

            # Calculate total market weight for the given rating and term bucket
            df[column_name] = df.apply(lambda row: ftse_data.loc[
                (ftse_data[column_to_look_in] == rating) &
                (ftse_data['TermPt'] < upper_b) &
                (ftse_data['TermPt'] >= lower_b) &
                (ftse_data['TermPt'] < row['Upper_Bound']) &
                (ftse_data['TermPt'] > row['Lower_Bound'] - 0.0001)
            ]['marketweight_noREITs'].sum(), axis=1)

            total_universe_weights.loc[rating, x + 1] = df[column_name].sum()
            
            # Dividing by the sum of the column to get the weight as a percentage of the subindex; ie,
            # Normalize by the sum of market weights
            df[column_name] = df[column_name] / df[column_name].sum()

        weight_dict[rating] = df

    return weight_dict, total_universe_weights


def create_general_shock_table() -> pd.DataFrame:
    """
    Creates a general shock table to calculate shocks for each security type.

    Returns:
    pd.DataFrame: A DataFrame containing the shock values for different term buckets.
    """
    shock_size = 0.0001
    buckets = [0, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30, 100]
    
    # Create a DataFrame for shocks with 70 term intervals
    df = pd.DataFrame(columns=buckets, index=list(np.linspace(start=0.5, stop=35, num=70)))
    df[0] = df.index  # Initialize the first column with term points
    
    # Apply shock formula for each bucket
    for i in range(1, 11):
        df[buckets[i]] = df.apply(lambda row: (
            (1 - (row[0] - buckets[i]) / (buckets[i + 1] - buckets[i])) * shock_size) 
            if (buckets[i] <= row[0] <= buckets[i + 1]) 
            else (((row[0] - buckets[i - 1]) / (buckets[i] - buckets[i - 1]) * shock_size) 
                  if buckets[i - 1] <= row[0] <= buckets[i] else 0), axis=1)

    # Drop the last bucket (100) as it is not needed
    df = df.drop(100, axis=1)
    
    return df


# Interpolating for half-years - side-eff 1

# Applies the shocks to the bond curves for each rating and store results in shocks_dict - side eff 2 (main eff...major purpose)
def create_shock_tables(curves): # CURVES are BOND CURVES LOL
    """
    Function to create up and down shock tables for bond curves
    """
    # makes a dictionary containing tables for up shocks and down shocks for each rating
    shocks_dict = {}
    up_shocks = create_general_shock_table()
    down_shocks = create_general_shock_table()
    down_shocks = -down_shocks
    down_shocks[0] = -down_shocks[0]

    #### Interpolates BOND CURVES for half years
    ## (*begin) Interpolates half-year curves (avg of 1, 2y rate for 11/12yr curve
        # 1+1/2 is average of 1y and 2y rate

    # Interpolating bond curves for half-year intervals
    curves_mod = pd.DataFrame(
        columns=['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate'],
        index=list(np.linspace(start=0.5, stop=35, num=70)))

    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for i in range(1, 35):
            curves_mod.loc[i, rating] = curves.loc[i, rating]
        for yr in np.linspace(start=1.5, stop=34.5, num=34):
            curves_mod.loc[yr, rating] = (curves.loc[yr + .5, rating] + curves.loc[yr - .5, rating]) / 2

    curves_mod.loc[0.5] = curves_mod.loc[1]
    curves_mod.fillna(method='ffill', inplace=True) ## (*end)

    ####

    # Apply up and down shocks to bond curves
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        for direction in ['Up', 'Down']:
            table_name = rating + ' - ' + direction
            if direction == 'Up':
                df = up_shocks
            else:
                df = down_shocks

            # Add shocks to bond curves
            df = df.add(curves_mod[rating], axis=0)
            df[0] = curves_mod[rating]

            # Apply power function to simulate the bond curve transformation after shocks
            df += 1
            df['Powers'] = df.index
            df = df.pow(df['Powers'], axis=0)
            df = 1 / df
            df = df.drop('Powers', axis=1)

            shocks_dict[table_name] = df

    return shocks_dict

# (*begin) Takes each year and looks at rating and FTSE universe (half-year would be from .25 to .75; up quarter year and down quarter year for half year, and so on for every year

# Function to calculate the average coupon rate for a specific bond rating and year
# It uses the FTSE data to filter bonds based on the given rating and term (maturity year).
# The average coupon is weighted by the market weight of the bond, excluding REITs.
def calc_avg_coupon(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the average coupon rate for a specific bond rating and year, weighted by the market weight of the bond.

    Parameters:
    year (float): The specific year (maturity) to calculate the coupon for.
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The average coupon rate for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) &
                       (ftse_data['marketweight_noREITs'] > 0)]

    # If no bonds match the criteria, return a coupon rate of 0
    if df.empty:
        return 0

    # Otherwise, calculate the weighted average coupon rate, dividing by 2 for semi-annual coupon payments. As follows:
        # 1. Multiply the market weight by the coupon rate and divide by the market value-adjusted interest (mvai).
        # 2. Divide the sum of these weighted values by the sum of market weights/mvai.
    avg_coupon = ((df['marketweight_noREITs'] * df['annualcouponrate'] / df['mvai']).sum() /
                  (df['marketweight_noREITs'] / df['mvai']).sum()) / 2  # Divide by 2 to account for semi-annual coupons
    
    # Return the calculated average coupon rate for the given rating and year (average was 0 if no matching bonds were found from FTSE bond databank)
    return avg_coupon



# Function to calculate the present value (PV) of bonds for a specific rating and year
# It uses the FTSE data to filter bonds based on the rating and term and then calculates the PV.
def calc_pv(year: float, rating: str, ftse_data: pd.DataFrame) -> float:
    """
    Calculates the present value (PV) of bonds for a specific bond rating and year.

    Parameters:
    year (float): The specific year (maturity) to calculate the PV for.
    rating (str): The bond rating category (e.g., 'Federal', 'CorporateAAA_AA', etc.).
    ftse_data (pd.DataFrame): A DataFrame containing FTSE bond data.

    Returns:
    float: The present value (PV) for the specified bond rating and year.
    """
    # Determine the column to filter by: "RatingBucket" for most bonds, or "Sector" for 'Corporate'
    column_to_look_in = "RatingBucket"
    if rating == 'Corporate':
        column_to_look_in = "Sector"  # Corporate bonds are filtered by 'Sector'

    # Define the term bounds (quarter-year before and after the specified year)
    lower_bound = year - 0.25
    upper_bound = year + 0.25

    # Filter FTSE data for bonds that:
    # 1. Match the rating or sector
    # 2. Have a term (maturity year) within the bounds
    # 3. Have a positive market weight excluding REITs
    df = ftse_data.loc[(ftse_data[column_to_look_in] == rating) &
                       (ftse_data['TermPt'] < upper_bound) &
                       (ftse_data['TermPt'] > (lower_bound - 0.001)) &
                       (ftse_data['marketweight_noREITs'] > 0)]

    # If no bonds match the criteria, return a PV of 0
    if df.empty:
        return 0

    # Otherwise, calculate the present value (PV) by summing up the product of the market weight and the bond's 
    # market value-adjusted interest (mvai), then dividing by the sum of the market weights.
    
    # I.e., Calculate the present value (PV) as the weighted sum of market value-adjusted interest (mvai)
    pv = (df['marketweight_noREITs'] * df['mvai']).sum() / df['marketweight_noREITs'].sum()

    # Return the calculated present value for the given rating and year
    return pv



### Begin

## USED FUNCTIONs ##

# (*end)

## USED FUNCTION ##

# Input: ftse_data - a DataFrame containing bond information.
# Output: cf_dict - a dictionary of cashflow tables and their respective present values for each bond rating.
def create_cf_tables(ftse_data: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        # uses the average coupon rate to calculate annual cashflows for each rating type

    """
    Creates cashflow tables for each bond rating based on FTSE data.

    This function calculates annual cashflows for different bond rating types based on the FTSE data.
    It creates a cashflow table for each rating and calculates the present value (PV) and coupon rates for each term bucket.
    
    Parameters:
    ftse_data (pd.DataFrame): A DataFrame containing bond information.

    Returns:
    Dict[str, pd.DataFrame]: A dictionary of cashflow tables and their respective present values for each bond rating.
    """
    cf_dict = {}  # Dictionary to store cashflow tables for each rating
    years = list(np.linspace(start=0.5, stop=35, num=70))  # 70 half-year intervals (from 0.5 to 35 years)
    buckets = list(np.linspace(start=0.5, stop=35, num=70))  # 70 term buckets (from 0.5 to 35 years)
    """
    Unecessary, as overwritten within the loop each time. Hence, removal of this commented-out area:
    # Create an empty DataFrame with columns for each year and rows for each bucket
    df = pd.DataFrame(columns=years, index=buckets)
    df.insert(0, 'Bucket', buckets)  # Add a 'Bucket' column representing the term
    df.insert(1, 'Principal', 100)  # Initialize with a default principal value of 100
    """
    # Iterate through each bond rating type to calculate cashflows
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        # Create a new DataFrame for the current rating
        df = pd.DataFrame(columns=years, index=buckets)
        df.insert(0, 'Bucket', buckets)  # Term buckets
        df.insert(1, 'Principal', 100)   # Default principal value of 100

        # Calculate present value (PV) and average coupon for each bucket
        df['PV'] = df.apply(lambda row: calc_pv(row['Bucket'], rating, ftse_data), axis=1)
        df['Coupon'] = df.apply(lambda row: calc_avg_coupon(row['Bucket'], rating, ftse_data), axis=1)

        # Move the coupon column to after 'Principal'
        coupons = df.pop(df.columns[-1])
        df.insert(2, 'Coupon', coupons)

        # Calculate cashflows for each term bucket based on coupon and principal
        for col in np.linspace(start=0.5, stop=35, num=70):
            df[col] = df.apply(lambda row: row['Coupon'] if row['Bucket'] > col
            else ((row['Coupon'] + row['Principal']) if row['Bucket'] == col else 0), axis=1)

        # Store the cashflow table and PV table for the rating
        cf_dict[rating] = df.iloc[:, :73]  # Cashflow table
        cf_dict[rating + 'PV'] = df.iloc[:, 73]  # Present value

    return cf_dict


# you can data transform it first
# THEN do the code
# do it fast then it'll be well

from typing import Dict

def create_sensitivity_tables(cashflows: Dict[str, pd.DataFrame], shocks: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Calculates cashflow sensitivities based on shocks applied to bond curves.

    Parameters:
    cashflows (Dict[str, pd.DataFrame]): A dictionary containing cashflow tables for each bond rating.
    shocks (Dict[str, pd.DataFrame]): A dictionary containing shock tables for each bond rating.

    Returns:
    Dict[str, pd.DataFrame]: A dictionary of sensitivity tables for each bond rating.
    """
    sensitivities_dict = {}  # Dictionary to store sensitivity tables
    buckets_krd = [0, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30]  # KRD buckets

    # Iterate through each bond rating type to calculate sensitivities
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        # Retrieve cashflows and shocks for the current rating
        cfs = cashflows[rating]  # Retrieve cashflows for the current rating
        ups = shocks[rating + ' - Up']  # Retrieve up shock table for the current rating
        downs = shocks[rating + ' - Down']  # Retrieve down shock table for the current rating

        """
        ## sumproduct for each, changed to get the sensitivities
        # cahsflows for the square (70*70) table, and it fits into the 10*70 sensitivities that it matches up to - sum to each one, cahsflow*shocks.
        """
        # Create empty DataFrames to store sensitivity data for up and down shocks
        df_up = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_up.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        df_down = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_down.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        # Calculate sensitivities by summing the product of cashflows and shocks
        for x in range(1, 11):
            for i in range(70):
                df_up.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * ups.iloc[:, x])  # Multiply cashflows by up shocks
                df_down.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * downs.iloc[:, x])  # Multiply cashflows by down shocks

        # Calculate the average sensitivity (difference between down and up shocks divided by 2)
        average_sensitivity = (df_down - df_up) / 2 * 10000

        # Add bucket information and transpose for better readability
        average_sensitivity['Bucket'] = list(np.linspace(start=.5, stop=35.5, num=71))
        average_sensitivity = average_sensitivity.transpose()
        average_sensitivity = average_sensitivity.drop(70, axis=1)
        average_sensitivity = average_sensitivity.iloc[1:]

        # Insert bucket names for KRD
        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        for x in range(10):
            for i in range(70):
                # Safe division, handling division by zero and inf
                numerator = average_sensitivity.iloc[x, i + 1]
                denominator = cashflows[rating + 'PV'].iloc[i]

                # Use np.divide with where clause to avoid division by zero and handle inf
                average_sensitivity.iloc[x, i + 1] = np.divide(numerator, denominator, out=np.zeros_like(numerator), where=denominator != 0)

        # Store the calculated sensitivity table for the rating
        sensitivities_dict[rating] = average_sensitivity

        """ Try this if the above doesn't work:
        # Insert bucket names for KRD and normalize by present values (PV)
        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        # Removed: avg_sensitivity = average_sensitivity - Brenda
        
        # Normalize the sensitivities by the cashflow present values (PV)
        for x in range(10):
            for i in range(70):
                average_sensitivity.iloc[x, i + 1] = average_sensitivity.iloc[x, i + 1] / cashflows[rating + 'PV'].iloc[i]

        # Store the calculated sensitivity table for the rating
        sensitivities_dict[rating] = average_sensitivity
        """

    return sensitivities_dict


"""    # Create empty DataFrames for storing up and down shock sensitivities
        df_up = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_up.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        df_down = pd.DataFrame(columns=buckets_krd[1:], index=range(71))
        df_down.insert(0, 'Bucket', list(np.linspace(start=0, stop=35, num=71)))

        # Old:
        #'''
        for x in range(1,11):
            for i in range(70):

                df_up.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * ups.iloc[:, x])  # cashflows table, sum prod
                df_down.iloc[i, x] = np.sum(cfs.iloc[i, 3:] * downs.iloc[:, x])
        #'''
        # New (Brenda):
        #df_up = np.sum(cfs.iloc[:, 3:].values * ups.values[:, 1:], axis=1)
        #df_down = np.sum(cfs.iloc[:, 3:].values * downs.values[:, 1:], axis=1)
        # New (Brenda)

        up_shock_sensitivities = df_up
        down_shock_sensitivities = df_down
        average_sensitivity = (down_shock_sensitivities - up_shock_sensitivities)/2 * 10000

        #average_sensitivity = (average_sensitivity.divide(ups.iloc[:,0]*100) * 10000).iloc[:,1:]
        average_sensitivity['Bucket'] = list(np.linspace(start=.5, stop=35.5, num=71))

        average_sensitivity = average_sensitivity.transpose()

        average_sensitivity = average_sensitivity.drop(70, axis=1)

        average_sensitivity = average_sensitivity.iloc[1:]

        average_sensitivity.insert(0, 'Bucket', [1, 2, 3, 5, 7, 10, 15, 20, 25, 30])

        avg_sensitivity = average_sensitivity

        for x in range(10):
            for i in range(70):
                average_sensitivity.iloc[x, i + 1] = avg_sensitivity.iloc[x, i + 1] / cashflows[rating + 'PV'].iloc[i]

        sensitivities_dict[rating] = average_sensitivity

    return sensitivities_dict"""
## (*end)

from typing import Dict

    # makes the final KRD table based on sensitivities and market weight and puts it all together in one dataframe
def make_krd_table(weights: Dict[str, pd.DataFrame], sensitivities: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Creates the final Key Rate Duration (KRD) table by combining market weights and cashflow sensitivities for each bond rating.

    Parameters:
    weights (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing market weights for each bond rating and maturity bucket.
    sensitivities (Dict[str, pd.DataFrame]): A dictionary of DataFrames containing sensitivities for each bond rating and maturity bucket.

    Returns:
    pd.DataFrame: A combined KRD table for all bond ratings and term buckets.
    """
    KRDs = {}
    cols = ['rating', 'term', 'bucket1', 'bucket2', 'bucket3', 'bucket4', 'bucket5', 'bucket6']
    buckets = [1, 2, 3, 5, 7, 10, 15, 20, 25, 30]

    # Iterate over each bond rating to calculate KRD values
    for rating in ['Federal', 'Provincial', 'CorporateAAA_AA', 'CorporateA', 'CorporateBBB', 'Corporate']:
        df = pd.DataFrame(columns=cols, index=range(10))
        df['term'] = buckets  # Assign bucket terms to the DataFrame
        df['rating'] = rating  # Set bond rating

        # Calculate KRD by multiplying sensitivities with market weights for each bucket
        for x in range(2, 8):
            df.iloc[:, x] = df.apply(lambda row: (
                sensitivities[rating].loc[sensitivities[rating]['Bucket'] == row['term']].iloc[:, 1:].values[0] *
                weights[rating].iloc[:, (x + 1)]
            ).sum(), axis=1)

        KRDs[rating] = df  # Store KRD DataFrame in the dictionary

    # Concatenate all rating-specific KRD DataFrames into one final DataFrame
    final_krd_df = pd.concat([KRDs['Federal'], KRDs['Provincial'], KRDs['CorporateAAA_AA'], 
                              KRDs['CorporateA'], KRDs['CorporateBBB'], KRDs['Corporate']], ignore_index=True)

    final_krd_df.fillna(0, inplace=True)  # Replace NaN values with 0
    return final_krd_df



import os
from datetime import datetime

def getBSPath(date: datetime) -> str:
    """
    Generates the file path for the Segmented Balance Sheet (SBS) file based on the provided date.

    Parameters:
    date (datetime): The date for which to retrieve the SBS file path.

    Returns:
    str: The file path for the Segmented Balance Sheet file.
    
    Raises:
    Exception: If the file path is not found.
    """
    bsFileName = 'SEGMENTED BALANCE SHEET-{year}.xlsx'
    # bsPathRoot = '\\\\estorage.equitable.int\\pcshared\\Financial Reporting\\Segmented Balance Sheets\\'
    bsPathRoot = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), 'Benchmarking')

    # Construct paths for the current year and the previous year
    current_year_path = os.path.join(bsPathRoot, bsFileName.format(year=date.strftime('%Y')))  # (*Brenda:) Removes the error if used on MacOS/Linux systems (uses OS.path.join over manual concatenation using Windows-assumed '\\')
    prev_year_path = os.path.join(bsPathRoot, bsFileName.format(year=str(date.year - 1)))

    # Check if the current year's file exists, else look in prior years' folder
    if os.path.exists(current_year_path):
        path = current_year_path
    else:
        path = os.path.join(bsPathRoot, 'Prior Years', date.strftime('%Y'), bsFileName.format(year=date.strftime('%Y')))

    # Raise an exception if the file path doesn't exist
    if not os.path.exists(path):
        raise Exception(f'Path not found: {path}')  # (*Brenda: Shifted error messaging to be more consistent formatting)

    return path



import pandas as pd
import numpy as np
from scipy import interpolate
import openpyxl

def get_expected_returns() -> pd.DataFrame:
    """
    Reads and interpolates expected bond returns from the "Parallel_tilt_curve_history.xlsx" file for various bond ratings and maturity terms.

    Returns:
    pd.DataFrame: A DataFrame containing the interpolated expected returns for different bond ratings and term assumptions.
    """
    file_name = "Parallel_tilt_curve_history.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    expected_returns = pd.DataFrame()
    ratings = ['Federal', 'Provincial', 'corporateAAA_AA', 'corporateA', 'corporateBBB']

    # Read the expected return data for each bond rating
    for sheet in ['analysis_quarterly_RF', 'analysis_quarterly_prov', 'analysis_quarterly_AA', 'analysis_quarterly_A', 'analysis_quarterly_BBB']:
        rownum = 27 if sheet == 'analysis_quarterly_RF' else 22
        ws = workbook[sheet]
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        returns = df.loc[rownum:rownum, 'term1':'term30']
        expected_returns = pd.concat([expected_returns, returns], ignore_index=True)

    # Assign bond ratings to the expected returns DataFrame
    expected_returns['ratings'] = ratings
    expected_returns.set_index('ratings', inplace=True)

    # Term assumptions for interpolation
    term_assumptions = [2, 7, 12, 17, 23, 29]
    return_assumptions = pd.DataFrame(columns=[0, 1, 2, 3, 4, 5])

    x = [1, 2, 3, 4, 5, 7, 10, 20, 30]
    # Interpolate expected returns for each rating across terms
    for rating in ratings:
        y = expected_returns.loc[rating].to_numpy()
        temp = interpolate.splrep(x, y, s=0)
        xnew = np.arange(1, 31)
        ynew = interpolate.splev(xnew, temp, der=0)
        return_assumptions.loc[rating] = ynew[term_assumptions]

    return return_assumptions / 100  # Convert to percentage returns


## New fUNCTION to read in stuff (no need from balance sheet) - check it easier from the balance sheet, easier
# helper function to read in from balance sheet (execel) not from the comapny blcsht its

# so have a function for reading in data
# for modifiability
# yu don't know how much i know the code - which is a lot
# i was doing it durin

# before it was reading in assets from balance sheet, but now its reading in data from excel sheet

# ReadInData
# this class also process it

# i forgot, bond class or calc omg
# recall and reconnect
#check the idea again
# thibk again for what makes sense///what was it again
# what did i think and mention


from datetime import datetime

def BSTotals(given_date: datetime, sheet_version: int) -> dict:
    """
    Retrieves the balance sheet totals from the "SBS Totals.xlsx" file based on the provided date.

    Parameters:
    given_date (datetime): The date for which the balance sheet totals are requested.
    sheet_version (int): Determines if totals or segments are returned (1 for segments, 0 for totals).

    Returns:
    dict: A dictionary containing balance sheet totals for different categories.
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    year_quarter = f"{year}Q{quarter}"

    file_name = "SBS Totals.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input, data_only=True)

    # Retrieve the appropriate worksheet based on sheet version
    ws = workbook[year_quarter] if sheet_version == 1 else workbook[year_quarter + ' (Total)']
    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # Extract and return totals for relevant categories
    totals = {
        'ACCUM': df.loc[2, 'ACCUM'],
        'PAYOUT': df.loc[2, 'PAYOUT'],
        'UNIVERSAL': df.loc[2, 'UNIVERSAL'],
        'NONPAR': df.loc[2, 'NONPAR'],
        'GROUP': df.loc[2, 'GROUP'],
        'PARCSM': df.loc[2, 'PARCSM'],
        'SEGFUNDS': df.loc[2, 'SEGFUNDS'],
        'Surplus': df.loc[2, 'Surplus'],
        'Total': df.loc[2, 'Total']
    } # rewrote BSTotals to reduce unecessary code / overwritten initializaitons

    return totals


def percents(given_date: datetime, curMonthBS: bool = False) -> pd.DataFrame:
    """
    Retrieves asset mix percentages from the "Mix_hardcoded.xlsx" file for the given date.

    Parameters:
    given_date (datetime): The date for which the asset mix percentages are requested.
    curMonthBS (bool): If True, adjusts the quarter to the next one if applicable. Default is False.

    Returns:
    pd.DataFrame: A DataFrame containing asset mix percentages for various bond ratings.
    """
    year = given_date.strftime("%Y")
    quarter = ((given_date.month - 1) // 3) + 1
    if curMonthBS and quarter < 4:
        quarter += 1
    year_quarter = year + "Q" + str(quarter)

    file_name = "Mix_hardcoded.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    workbook = openpyxl.load_workbook(path_input)
    ws = workbook[year_quarter]  # i.e., ws = workbook['2024Q1']

    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns).set_index('rating')
    
    # Initialize surplus and SEGFUNDS columns
    df['Surplus'] = df['SEGFUNDS'] = 0

    # Filter rows to include only relevant bond categories
    df = df.loc[['Public Federal', 'Public Provincial', 'Public Corporate - AA', 'Public Corporate - A', 
                 'Public Corporate - BBB', 'MortgagesInsured', 'MortgagesConv', 
                 'PrivateAA', 'PrivateA', 'PrivateBBB', 'PrivateBB_B']]

    # Rename columns for readability
    df.rename({'Public Federal': 'Federal', 'Public Provincial': 'Provincial', 
               'Public Corporate - AA': 'CorpAAA_AA', 'Public Corporate - A': 'CorpA', 
               'Public Corporate - BBB': 'CorpBBB'}, inplace=True)
    
    return df

from typing import Dict

def solution_dollar_amounts(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculates the dollar allocation for each bond rating across different portfolios based on the asset mix and solution data.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations by rating.

    Returns:
    pd.DataFrame: A DataFrame with dollar allocations for each portfolio and rating.
    """
    weights = Asset_mix[['Accum', 'group', 'ul', 'Payout', 'np']].stack().sort_index()  # (*Brenda)
    weights2 = weights.reset_index(drop=True)
    
    # Filter the solution DataFrame to exclude 'Liability' and 'Total' portfolios
    sols = solution_df[(solution_df['portfolio'] != 'Liability') & (solution_df['portfolio'] != 'Total')].set_index(['rating', 'portfolio']).sort_index()  # Python has great order-of-operations (*Brenda)
    sols2 = sols.reset_index(drop=True)
    
    # Calculate weighted dollar allocations
    w = sols2.mul(weights2, axis=0)
    w['rating'] = sols.reset_index()['rating']
    w['portfolio'] = sols.reset_index()['portfolio']
    w = w.set_index(['portfolio', 'rating'])

    # Group by rating and sum for total allocation
    w_grouped = w.groupby('rating')
    for index, row in w_grouped:
        total_values = row.sum()
        total_values['rating'] = index
        total_values['portfolio'] = 'Total'
        w = pd.concat([w, pd.DataFrame(total_values).T.set_index(['portfolio', 'rating'])])

    return w.reset_index()

''' This function takes in the asset mix and the solved solution up to this point to calculate how much of the total allocation has been allocated in each portfolio. Those weights are used as bounds for the total optimization''' # Old comment vs my interpretation (kept both in case err - for now)
def get_bnds_for_total(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculates the bounds for total optimization based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the upper and lower bounds for optimization for each rating.
    """
    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')
    dollars = Asset_mix['Total']
    
    # Calculate bounds by dividing total allocation by asset mix
    bounds = total.div(dollars, axis=0)
    bounds = bounds.where(bounds > 0, 0)
    
    return bounds


def liabilities_table(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a table of liability allocations for each bond rating based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the liability allocations by bond rating.
    """
    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')
    dollars = total.sum(axis=1)
    
    # Calculate liability allocations by dividing by total dollars
    liabilities = total.div(dollars, axis=0)
    liabilities['rating'] = liabilities.index
    liabilities['portfolio'] = 'Liability'
    
    return liabilities.reset_index(drop=True)


def surplus_table(Asset_mix: pd.DataFrame, solution_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a table of surplus allocations for each bond rating based on the asset mix and portfolio allocations.

    Parameters:
    Asset_mix (pd.DataFrame): A DataFrame containing the asset mix weights for different portfolios.
    solution_df (pd.DataFrame): A DataFrame containing the solved portfolio allocations.

    Returns:
    pd.DataFrame: A DataFrame containing the surplus allocations by bond rating.
    """
    npt_weights = Asset_mix['Total']
    npt_sol = solution_df[solution_df['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')

    # Calculate optimized solution using weights
    optimization_sol = npt_sol.mul(npt_weights, axis=0)

    sol = solution_dollar_amounts(Asset_mix, solution_df)
    total = sol[sol['portfolio'] == 'Total'].drop(columns=['portfolio']).set_index('rating')

    # Calculate the surplus by subtracting total from optimization
    total = optimization_sol - total
    dollars = total.sum(axis=1)
    
    # Calculate surplus allocations by dividing by total dollars
    surplus = total.div(dollars, axis=0)
    surplus['rating'] = surplus.index
    surplus['portfolio'] = 'Surplus'
    
    return surplus.reset_index(drop=True)


from typing import List
from datetime import datetime

def calc_bounds(given_date: datetime, portfolio: str, total: float) -> List[List[float]]:
    """
    Calculates the optimization bounds for cashflow buckets based on historical data for a given portfolio.

    Parameters:
    given_date (datetime): The date to base the historical data on.
    portfolio (str): The portfolio type (such as 'ul' or 'np') for which bounds are being calculated.
    total (float): The total portfolio value to normalize bounds.

    Returns:
    List[List[float]]: A list of bounds for each cashflow bucket, defining the lower and upper limits.
    """
    if portfolio not in ['ul', 'np']:
        return [[0, 1]] * 6  # Default bounds for other portfolios

    year = given_date.strftime('%Y')
    year_folder = given_date.strftime('%Y')
    quarter = ((given_date.month - 1) // 3) + 1
    prev_quarter = quarter - 1
    if prev_quarter == 0:
        prev_quarter = 4
        year = str(given_date.year - 1)

    quarter = str(quarter)
    prev_quarter = str(prev_quarter)

    # Construct the file name based on the quarter and year
    if given_date.year == 2024 and quarter == '1':
        file_name = f"{portfolio} IFE Estimate Q1 2024.xlsx"
    else:
        file_name = f"{portfolio} IFE Estimate Q{quarter} {year}.xlsx"

    path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation", year_folder, 
                              'IFE estimates', f'Q{quarter}', file_name)
    try:
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)
    except FileNotFoundError:
        file_name = f"{portfolio} IFE Estimate Q{prev_quarter} {year} to Q{quarter}.xlsx"
        path_input = os.path.join(sysenv.get('LOB_MANAGEMENT_DIR'), "Investment Income Explanation", 
                                  year_folder, 'IFE estimates', f'Q{quarter}', file_name)
        workbook = openpyxl.load_workbook(path_input, data_only=True, read_only=True)

    ws = workbook['CF']
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    # Retrieve the present value (PV) for each cashflow bucket
    cf_pvs = df.iloc[1:7, 34].tolist()
    bounds = []

    # Define bounds for each bucket, allowing short positions if PV is negative
    for pv in cf_pvs:
        if pv >= 0:
            bounds.append([0, 6])
        else:
            bounds.append([pv / total, 6])

    return bounds

''' given a df with a multi-index, portfolio and rating, this function will sum all rows with the same rating, and append the sum to a new row with portfolio 'Total' '''
def get_totals_for_rating(df: pd.DataFrame, reset_index: bool = False) -> pd.DataFrame:
    """
    Summarizes the total values for each rating in the given DataFrame by aggregating portfolios.

    Parameters:
    df (pd.DataFrame): A DataFrame with multi-index of 'portfolio' and 'rating'.
    reset_index (bool): Whether to reset the index after aggregating. Default is False.

    Returns:
    pd.DataFrame: A DataFrame with total values for each rating, with an additional row for portfolio 'Total'.
    """
    print(df)

    df_copy = df.copy()
    df_grouped = df_copy.groupby('rating')

    # Sum all rows with the same rating and append the total row
    for index, row in df_grouped:
        total_values = row.sum()
        total_values['rating'] = index
        total_values['portfolio'] = 'Total'
        total_values_df = pd.DataFrame(total_values).T.set_index(['portfolio', 'rating'])
        df_copy = pd.concat([df_copy, total_values_df])

    return df_copy.reset_index() if reset_index else df_copy


def public_sensitivities() -> pd.DataFrame:
    """
    Retrieves public asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for public asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    sheet = 'public'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df


def private_sensitivities() -> pd.DataFrame:
    """
    Retrieves private asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for private asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    sheet = 'private'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df


def mortgage_sensitivities() -> pd.DataFrame:
    """
    Retrieves mortgage asset class sensitivities from the "Targets By Asset Class.xlsx" file.

    Returns:
    pd.DataFrame: A DataFrame containing the sensitivities for mortgage asset classes.
    """
    file_name = "Targets By Asset Class.xlsx"
    path_input = os.path.join(sysenv.get('PORTFOLIO_ATTRIBUTION_DIR'), "Benchmarking", file_name)
    sheet = 'mortgage'
    workbook = openpyxl.load_workbook(path_input, data_only=True)
    ws = workbook[sheet]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    return df
